import os
import sys
# === Linux X11 fix: disable AT-SPI accessibility bridge ===
# Must be set BEFORE any Qt imports.
# On ALT Linux / RHEL-like systems with X11, Qt6 accessibility bridge (AT-SPI)
# crashes in native C++ during window.show(), causing a segfault.
# QOpenGLContext().isValid() == False confirms software-only rendering environment.
if sys.platform == "linux":
    os.environ.setdefault("QT_ACCESSIBILITY", "0")
    os.environ.setdefault("QT_LINUX_ACCESSIBILITY_ALWAYS_ON", "0")
    os.environ.setdefault("QT_XCB_NO_MITSHM", "1")
import re
from pathlib import Path
from typing import List, Optional, Tuple, Any
import json
import subprocess
import datetime

import pandas as pd
import openpyxl
from openpyxl.utils import get_column_letter

from PySide6.QtWidgets import (
    QApplication, QMainWindow, QVBoxLayout, QHBoxLayout, QWidget,
    QPushButton, QLineEdit, QLabel, QFileDialog, QComboBox, QTextEdit,
    QTableWidget, QTableWidgetItem, QCheckBox, QMessageBox,
    QDialog, QScrollArea, QSizePolicy, QInputDialog
)
from PySide6.QtGui import QIntValidator

# ==================== КОНСТАНТЫ ====================
GLOBAL_CONFIG_FILE = "global_config.json"

LAST_USED_PRESET_KEY = "last_used_preset"
PRESETS_KEY = "presets"

SOURCE_FILES_KEY = "source_file_paths"
OUTPUT_DIR_KEY = "output_directory"
OUTPUT_FILENAME_KEY = "output_filename"
LAST_SELECTED_SHEET_KEY = "selected_sheet_name"
MARKER_TEXT_KEY = "marker_text"
MARKER_ROW_KEY = "marker_row"
MARKER_COL_KEY = "marker_col"
USE_CELL_COORD_KEY = "use_cell_coordinates_for_marker"
NUM_COLUMNS_KEY = "num_columns"
INSERT_HEADER_KEY = "insert_header"
FILTER_OPTION_KEY = "filter_option"
NUMERIC_COLUMNS_KEY = "numeric_columns"

OUTPUT_SHEET_NAME = "Combined_Data"

# ==================== ФУНКЦИИ РАБОТЫ С ФАЙЛАМИ ====================
def get_excel_files(file_paths: List[str]) -> List[str]:
    existing_files = [f for f in file_paths if os.path.exists(f) and (f.endswith('.xlsx') or f.endswith('.xlsm'))]
    return sorted(list(set(existing_files)))


# ==================== PresetManager ====================
class PresetManager:
    def __init__(self):
        self.global_config_path = Path(__file__).parent / GLOBAL_CONFIG_FILE
        self.global_config: dict = self._load_global_config()
        self.last_used_preset_name: Optional[str] = self.global_config.get(LAST_USED_PRESET_KEY)

    def _load_global_config(self) -> dict:
        if self.global_config_path.exists():
            try:
                with open(self.global_config_path, "r", encoding="utf-8") as f:
                    return json.load(f)
            except Exception as e:
                print(f"Ошибка при чтении {GLOBAL_CONFIG_FILE}: {e}")
        return {PRESETS_KEY: {}}

    def _save_global_config(self):
        try:
            with open(self.global_config_path, "w", encoding="utf-8") as f:
                json.dump(self.global_config, f, ensure_ascii=False, indent=4)
        except Exception as e:
            print(f"Ошибка при записи {GLOBAL_CONFIG_FILE}: {e}")

    def get_all_preset_names(self) -> List[str]:
        return sorted(list(self.global_config.get(PRESETS_KEY, {}).keys()))

    def save_preset(self, name: str, settings: dict):
        name = re.sub(r'[\W_]', '_', name)
        self.global_config.setdefault(PRESETS_KEY, {})[name] = settings
        self.last_used_preset_name = name
        self.global_config[LAST_USED_PRESET_KEY] = name
        self._save_global_config()
        return True

    def load_preset(self, name: str) -> Optional[dict]:
        presets = self.global_config.get(PRESETS_KEY, {})
        if name in presets:
            settings = presets[name]
            self.last_used_preset_name = name
            self.global_config[LAST_USED_PRESET_KEY] = name
            self._save_global_config()
            return settings
        return None

    def delete_preset(self, name: str):
        presets = self.global_config.get(PRESETS_KEY, {})
        if name in presets:
            del presets[name]
            if self.last_used_preset_name == name:
                self.last_used_preset_name = None
                self.global_config.pop(LAST_USED_PRESET_KEY, None)
            self._save_global_config()
            return True
        return False

# ==================== ФУНКЦИИ ОБРАБОТКИ ДАННЫХ ====================
def find_text_position_in_df(df: pd.DataFrame, search_text: str) -> Optional[Tuple[int, int]]:
    search_lower = search_text.lower()
    for row_idx, row in df.iterrows():
        for col_idx, cell_value in row.items():
            if str(cell_value).strip().lower() == search_lower:
                return (row_idx, col_idx)
    return None

def extract_data_from_file(
    file_path: str,
    sheet_name: str,
    marker_text: str,
    marker_row: int,
    marker_col: int,
    use_cell_coordinates_for_marker: bool,
    num_columns: int,
    insert_header: bool,
    header_row_values,  # передаётся по ссылке, обновляется при первом файле
    numeric_columns: Optional[List[int]] = None,   # относительные индексы столбцов (0-based)
    log_func=None
) -> Optional[pd.DataFrame]:
    filename = os.path.basename(file_path)
    log_prefix = f" [{filename}] " if log_func else ""

    def log(message):
        if log_func:
            log_func(log_prefix + message)
        else:
            print(log_prefix + message)

    # Проверка существования листа
    excel_file = pd.ExcelFile(file_path)
    if sheet_name not in excel_file.sheet_names:
        log(f"⚠ Лист '{sheet_name}' не найден.")
        return None, header_row_values

    # Определение стартовой позиции
    if use_cell_coordinates_for_marker:
        if marker_row < 0 or marker_col < 0:
            log(f"❌ Некорректные координаты ячейки: строка {marker_row+1}, столбец {marker_col+1}.")
            return None, header_row_values
        start_row = marker_row
        start_col = marker_col
        log(f"✓ Данные будут извлекаться, начиная со строки {start_row+1}, столбца {start_col+1} (по указанным координатам).")
    else:
        # Читаем весь лист как текст для поиска маркера
        df_full = pd.read_excel(file_path, sheet_name=sheet_name, header=None, dtype=str, keep_default_na=False)
        position = find_text_position_in_df(df_full, marker_text)
        if position is None:
            log(f"⚠ Текст маркера '{marker_text}' не найден.")
            return None, header_row_values
        start_row, start_col = position
        log(f"✓ Маркер '{marker_text}' найден в строке {start_row+1}, столбце {start_col+1}.")

    # Читаем данные, пропуская строки до маркера
    df = pd.read_excel(
        file_path,
        sheet_name=sheet_name,
        header=None,
        skiprows=start_row,
        dtype=str,
        keep_default_na=False
    )

    if df.empty:
        log("⚠ DataFrame пуст после пропуска строк.")
        return None, header_row_values

    if start_col >= len(df.columns):
        log(f"❌ Ошибка: стартовый столбец {start_col+1} вне диапазона ({len(df.columns)} столбцов).")
        return None, header_row_values

    end_col = start_col + num_columns
    selected_df = df.iloc[:, start_col:end_col].copy()

    # Приводим количество столбцов к заданному num_columns
    if selected_df.shape[1] > num_columns:
        selected_df = selected_df.iloc[:, start_col:end_col]          # обрезаем лишние
    elif selected_df.shape[1] < num_columns:
        for i in range(start_col +selected_df.shape[1], end_col):
            selected_df[i] = ""                                   # добавляем пустые столбцы
    # Теперь количество столбцов точно равно num_columns
    selected_df.columns = range(num_columns)

    # Обработка заголовка
    if insert_header:
        # Если заголовок ещё не был получен (первый файл), извлекаем его из первой строки
        if header_row_values is None:
            # Берём все значения первой строки (она содержит заголовки)
            header_row_values = selected_df.iloc[0].tolist()
            log(f"✓ Заголовок из первого файла: {header_row_values}")
        # Удаляем первую строку из данных, так как это заголовок
        if len(selected_df) > 1:
            selected_df = selected_df.iloc[1:]
        else:
            # Если после маркера только одна строка (заголовок), данных нет
            log("⚠ В файле только заголовок, нет данных для извлечения.")
            return None, header_row_values
    else:
        # Если заголовок не вставляется, первую удаляем
        selected_df = selected_df.iloc[1:]

    # Преобразование указанных столбцов в числовой тип
    if numeric_columns:
        for rel_col in numeric_columns:
            if rel_col < selected_df.shape[1]:
                try:
                    selected_df[rel_col] = pd.to_numeric(selected_df[rel_col], errors='coerce')
                    log(f"✓ Столбец {rel_col+1} (относительно начала) преобразован в числовой тип.")
                except Exception as e:
                    log(f"⚠ Ошибка преобразования столбца {rel_col+1} в число: {e}")

    # Добавляем столбец с именем исходного файла
    selected_df["Source_File"] = filename
    log(f"✓ Извлечено {len(selected_df)} строк.")
    return selected_df, header_row_values

def filter_dataframe(df: pd.DataFrame, filter_option: str, log_func=None) -> pd.DataFrame:
    def log(message):
        if log_func:
            log_func(message)
        else:
            print(message)

    if filter_option == "Удалить строки с пустым ПЕРВЫМ столбцом":
        if 0 in df.columns:
            initial_len = len(df)
            df = df.dropna(subset=[0])
            df = df[df[0].astype(str).str.strip() != ""]
            removed = initial_len - len(df)
            log(f"✓ Удалено {removed} строк с пустым первым столбцом.")
        else:
            log("⚠ Первый столбец не найден (индекс 0), фильтрация пропущена.")
    elif filter_option == "Удалить ПОЛНОСТЬЮ пустые строки":
        data_columns = [col for col in df.columns if col != "Source_File"]
        if data_columns:
            initial_len = len(df)
            df = remove_fully_empty_rows(df, data_columns, log_func)
            removed = initial_len - len(df)
            log(f"✓ Удалено {removed} полностью пустых строк.")
        else:
            log("⚠ Нет столбцов данных для проверки на пустоту, фильтрация пропущена.")
    else:
        log("✓ Фильтрация пропущена, сохранены все строки.")
    return df

def remove_fully_empty_rows(df, data_columns=None, log_func=None):
    """
    Удаляет строки, полностью пустые в указанных столбцах.
    
    Параметры:
        df : pandas.DataFrame
            Исходный DataFrame.
        data_columns : list, optional
            Список столбцов для проверки. Если None, используются все столбцы,
            кроме 'Source_File'.
        log_func : callable, optional
            Функция для логирования (например, self._log).
    
    Возвращает:
        pandas.DataFrame
            DataFrame без полностью пустых строк (индексы сброшены).
    """
    if data_columns is None:
        data_columns = [col for col in df.columns if col != "Source_File"]
    
    if not data_columns:
        if log_func:
            log_func("⚠ Нет столбцов для проверки на пустоту.")
        return df

    initial_len = len(df)

    # 1. Заменяем NaN на пустые строки
    # 2. Преобразуем всё в строки (на всякий случай)
    # 3. Удаляем пробелы по краям
    temp = df[data_columns].fillna('').astype(str).apply(lambda x: x.str.strip())

    # Строка считается пустой, если все её значения – пустые строки
    empty_mask = (temp == '').all(axis=1)

    # Удаляем пустые строки
    df_cleaned = df[~empty_mask].copy()
    df_cleaned.reset_index(drop=True, inplace=True)

    removed = initial_len - len(df_cleaned)
    if log_func:
        log_func(f"✓ Удалено {removed} полностью пустых строк.")
    return df_cleaned

def apply_auto_filter(file_path: str, sheet_name: str, header_row_index: int, log_func=None):
    def log(message):
        if log_func:
            log_func(message)
        else:
            print(message)

    try:
        workbook = openpyxl.load_workbook(file_path)
        sheet = workbook[sheet_name]

        max_row = sheet.max_row
        max_col_letter = get_column_letter(sheet.max_column)

        if max_row < header_row_index:
            log(f"⚠ Лист '{sheet_name}' содержит меньше строк ({max_row}) чем строка заголовка ({header_row_index}). Автофильтр не применен.")
            return

        filter_range = f"A{header_row_index}:{max_col_letter}{max_row}"
        sheet.auto_filter.ref = filter_range

        workbook.save(file_path)
        log(f"✓ Автофильтр применен к листу '{sheet_name}' на диапазоне {filter_range}.")
    except Exception as e:
        log(f"❌ Ошибка при применении автофильтра к файлу '{file_path}' на листе '{sheet_name}': {e}")

# ==================== ДИАЛОГ ПРЕДПРОСМОТРА ====================
class PreviewDialog(QDialog):
    def __init__(self, excel_file_path: str, sheet_name: str,
                 current_marker_row: int, current_marker_col: int,
                 current_numeric_columns_rel: List[int],
                 marker_text: str = "",
                 parent=None):
        super().__init__(parent)
        self.setWindowTitle(f"Предпросмотр: {os.path.basename(excel_file_path)} - {sheet_name}")
        self.setGeometry(100, 100, 1000, 700)

        self.selected_cell_text = ""
        self.selected_row = -1
        self.selected_col = -1
        self.initial_marker_row = current_marker_row
        self.initial_marker_col = current_marker_col
        self.marker_text = marker_text if marker_text is not None else ""
        self.numeric_columns_rel = current_numeric_columns_rel.copy()

        if current_marker_col != -1:
            self.numeric_columns_abs = [col + current_marker_col for col in self.numeric_columns_rel]
        else:
            self.numeric_columns_abs = self.numeric_columns_rel.copy()

        layout = QVBoxLayout(self)

        self.info_label = QLabel("Кликните на ячейку, чтобы выбрать текст маркера. Кликните на заголовок столбца, чтобы пометить его как числовой (отмечен *).")
        layout.addWidget(self.info_label)

        self.table_widget = QTableWidget()
        self.table_widget.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table_widget.setSelectionBehavior(QTableWidget.SelectItems)
        self.table_widget.setSelectionMode(QTableWidget.SingleSelection)
        self.table_widget.itemClicked.connect(self._item_clicked)
        layout.addWidget(self.table_widget)

        self.table_widget.horizontalHeader().sectionClicked.connect(self._header_clicked)

        self.selected_text_label = QLabel("Выбранный текст: ")
        layout.addWidget(self.selected_text_label)

        self.numeric_label = QLabel("")
        layout.addWidget(self.numeric_label)

        button_layout = QHBoxLayout()
        self.cancel_button = QPushButton("Отмена")
        self.cancel_button.clicked.connect(self.reject)
        self.confirm_button = QPushButton("Подтвердить выбор")
        self.confirm_button.clicked.connect(self.accept)
        self.confirm_button.setEnabled(False)
        button_layout.addWidget(self.cancel_button)
        button_layout.addWidget(self.confirm_button)
        layout.addLayout(button_layout)

        self._load_data(excel_file_path, sheet_name)
        self._try_auto_select_marker()
        self._preselect_initial_cell()
        self._update_header_stars()

    def _load_data(self, excel_file_path: str, sheet_name: str):
        try:
            df = pd.read_excel(excel_file_path, sheet_name=sheet_name, header=None, nrows=200, dtype=str, keep_default_na=False)
            if df.empty:
                self.table_widget.setRowCount(0)
                self.table_widget.setColumnCount(0)
                self.info_label.setText("Лист не содержит данных.")
                return

            self.table_widget.setRowCount(df.shape[0])
            self.table_widget.setColumnCount(df.shape[1])

            self.col_labels = []
            for c_idx in range(df.shape[1]):
                col_label = ""
                temp_c_idx = c_idx
                while temp_c_idx >= 0:
                    col_label = chr(65 + temp_c_idx % 26) + col_label
                    temp_c_idx = temp_c_idx // 26 - 1
                self.col_labels.append(col_label)

            self.table_widget.setHorizontalHeaderLabels(self.col_labels)
            self.table_widget.setVerticalHeaderLabels([str(i + 1) for i in range(df.shape[0])])

            for r_idx in range(df.shape[0]):
                for c_idx in range(df.shape[1]):
                    value = df.iat[r_idx, c_idx]
                    item = QTableWidgetItem(str(value) if value != "" else "")
                    self.table_widget.setItem(r_idx, c_idx, item)

            self.table_widget.resizeColumnsToContents()
            self.table_widget.horizontalHeader().setStretchLastSection(True)

            self.info_label.setText(f"Кликните на ячейку, чтобы выбрать текст маркера. Кликните на заголовок столбца, чтобы пометить его как числовой (отмечен *). Показаны первые {df.shape[0]} строк.")
        except Exception as e:
            QMessageBox.critical(self, "Ошибка загрузки", f"Не удалось загрузить данные листа: {e}")
            self.confirm_button.setEnabled(False)
            self.info_label.setText("Ошибка загрузки данных.")

    def _try_auto_select_marker(self):
        if self.initial_marker_col == -1 and self.marker_text and self.table_widget.rowCount() > 0:
            search_text = self.marker_text.strip().lower()
            for r in range(self.table_widget.rowCount()):
                for c in range(self.table_widget.columnCount()):
                    item = self.table_widget.item(r, c)
                    if item and item.text().strip().lower() == search_text:
                        self.table_widget.setCurrentCell(r, c)
                        self._item_clicked(item)
                        return

    def _preselect_initial_cell(self):
        if (self.initial_marker_row >= 0 and self.initial_marker_row < self.table_widget.rowCount() and
            self.initial_marker_col >= 0 and self.initial_marker_col < self.table_widget.columnCount()):
            self.table_widget.setCurrentCell(self.initial_marker_row, self.initial_marker_col)
            item = self.table_widget.item(self.initial_marker_row, self.initial_marker_col)
            if item:
                self._item_clicked(item)

    def _item_clicked(self, item: QTableWidgetItem):
        self.selected_row = item.row()
        self.selected_col = item.column()
        self.selected_cell_text = item.text().strip()
        self.selected_text_label.setText(f"Выбранный текст: '{self.selected_cell_text}' "
                                         f"(строка {self.selected_row + 1}, столбец {self.selected_col + 1})")
        self.confirm_button.setEnabled(True)
        self._recalc_abs_from_rel()
        self._update_header_stars()

    def _recalc_abs_from_rel(self):
        if self.selected_col != -1:
            self.numeric_columns_abs = [col + self.selected_col for col in self.numeric_columns_rel]
        else:
            self.numeric_columns_abs = self.numeric_columns_rel.copy()

    def _update_header_stars(self):
        for col in range(self.table_widget.columnCount()):
            current_text = self.col_labels[col].rstrip('*')
            if col in self.numeric_columns_abs:
                new_text = current_text + '*'
            else:
                new_text = current_text
            if self.col_labels[col] != new_text:
                self.col_labels[col] = new_text
                item = self.table_widget.horizontalHeaderItem(col)
                if item:
                    item.setText(new_text)
        if self.numeric_columns_abs:
            self.numeric_label.setText(f"Числовые столбцы (номера в Excel): {', '.join(str(i+1) for i in sorted(self.numeric_columns_abs))}")
        else:
            self.numeric_label.setText("Числовые столбцы: не выбраны")

    def _header_clicked(self, logical_index: int):
        if logical_index in self.numeric_columns_abs:
            self.numeric_columns_abs.remove(logical_index)
        else:
            self.numeric_columns_abs.append(logical_index)
        self._update_header_stars()

    def accept(self):
        if self.selected_col != -1:
            self.numeric_columns_rel = [idx - self.selected_col for idx in self.numeric_columns_abs if idx >= self.selected_col]
        else:
            self.numeric_columns_rel = self.numeric_columns_abs.copy()
        super().accept()

# ==================== ГЛАВНОЕ ОКНО ====================
class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Объединение Excel Файлов")
        self.setGeometry(100, 100, 800, 950)

        self._is_applying_preset: bool = False
        self.last_output_file_path: Optional[str] = None
        self.preset_manager = PresetManager()

        self.source_file_paths: List[str] = []
        self.output_directory: str = str(Path(__file__).parent.resolve())
        self.output_filename: str = "output.xlsx"
        self.selected_sheet_name: Optional[str] = None
        self.marker_text: str = ""
        self.marker_row: int = -1
        self.marker_col: int = -1
        self.use_cell_coordinates_for_marker: bool = False
        self.num_columns: int = 5
        self.insert_header: bool = True
        self.filter_option: str = "Без фильтрации"
        self.numeric_columns: List[int] = []

        self.first_file_path: Optional[str] = None

        self._init_ui()
        self._load_last_used_preset()

    def _init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout(central_widget)

        # Секция 0: Управление пресетами
        main_layout.addWidget(QLabel("<h2>0. Управление пресетами настроек</h2>"))
        preset_layout = QHBoxLayout()
        preset_layout.addWidget(QLabel("Текущий пресет:"))
        self.preset_combo = QComboBox()
        self.preset_combo.currentIndexChanged.connect(self._preset_selected)
        preset_layout.addWidget(self.preset_combo)
        self.btn_save_preset = QPushButton("Сохранить пресет")
        self.btn_save_preset.clicked.connect(self._save_current_preset)
        preset_layout.addWidget(self.btn_save_preset)
        self.btn_delete_preset = QPushButton("Удалить пресет")
        self.btn_delete_preset.clicked.connect(self._delete_current_preset)
        preset_layout.addWidget(self.btn_delete_preset)
        main_layout.addLayout(preset_layout)
        main_layout.addStretch(1)

        # Секция 1: Выбор файлов
        main_layout.addWidget(QLabel("<h2>1. Выберите исходные файлы (.xlsx, .xlsm)</h2>"))
        source_files_layout = QVBoxLayout()
        button_layout = QHBoxLayout()
        self.btn_select_source_files = QPushButton("Выбрать/Добавить файлы")
        self.btn_select_source_files.clicked.connect(self._select_source_files)
        button_layout.addWidget(self.btn_select_source_files)
        self.btn_clear_source_files = QPushButton("Очистить список файлов")
        self.btn_clear_source_files.clicked.connect(self._clear_source_files)
        button_layout.addWidget(self.btn_clear_source_files)
        source_files_layout.addLayout(button_layout)
        self.file_list_label = QLabel("<b>Выбранные файлы:</b>")
        source_files_layout.addWidget(self.file_list_label)
        self.file_list_text = QTextEdit()
        self.file_list_text.setReadOnly(True)
        self.file_list_text.setFixedHeight(120)
        source_files_layout.addWidget(self.file_list_text)
        main_layout.addLayout(source_files_layout)

        # Секция 2: Настройка параметров извлечения
        main_layout.addWidget(QLabel("<h2>2. Настройка параметров извлечения данных</h2>"))
        sheet_layout = QHBoxLayout()
        sheet_layout.addWidget(QLabel("Лист для обработки (из первого файла):"))
        self.sheet_combo = QComboBox()
        self.sheet_combo.currentIndexChanged.connect(self._sheet_selected)
        sheet_layout.addWidget(self.sheet_combo)
        main_layout.addLayout(sheet_layout)

        self.use_coords_checkbox = QCheckBox("Использовать координаты ячейки")
        self.use_coords_checkbox.setChecked(self.use_cell_coordinates_for_marker)
        self.use_coords_checkbox.stateChanged.connect(self._coords_checkbox_state_changed)
        main_layout.addWidget(self.use_coords_checkbox)

        marker_layout = QHBoxLayout()
        self.btn_preview_marker = QPushButton("Предпросмотр и выбор маркера/координат")
        self.btn_preview_marker.clicked.connect(self._open_preview_dialog)
        self.marker_text_edit = QLineEdit(self.marker_text)
        self.marker_text_edit.setPlaceholderText("Текст маркера (например, 'ID драйвера')")
        self.marker_text_edit.setReadOnly(True)
        self.marker_coords_label = QLabel("Координаты (строка, столбец):")
        self.marker_coords_edit = QLineEdit(f"({self.marker_row+1}, {self.marker_col+1})")
        self.marker_coords_edit.setPlaceholderText("(Строка, Столбец)")
        self.marker_coords_edit.setReadOnly(True)
        marker_layout.addWidget(self.btn_preview_marker)
        marker_layout.addWidget(self.marker_text_edit)
        marker_layout.addWidget(self.marker_coords_label)
        marker_layout.addWidget(self.marker_coords_edit)
        main_layout.addLayout(marker_layout)

        cols_layout = QHBoxLayout()
        cols_layout.addWidget(QLabel("Количество столбцов для извлечения:"))
        self.num_cols_edit = QLineEdit(str(self.num_columns))
        self.num_cols_edit.setValidator(QIntValidator(1, 100))
        self.num_cols_edit.textChanged.connect(self._save_current_settings_to_temp_preset)
        cols_layout.addWidget(self.num_cols_edit)
        main_layout.addLayout(cols_layout)

        self.insert_header_checkbox = QCheckBox("Вставлять строку-заголовок (из файла) в итоговый файл?")
        self.insert_header_checkbox.setChecked(self.insert_header)
        self.insert_header_checkbox.stateChanged.connect(self._save_current_settings_to_temp_preset)
        main_layout.addWidget(self.insert_header_checkbox)

        self.numeric_label = QLabel("Числовые столбцы: не выбраны")
        main_layout.addWidget(self.numeric_label)

        # Секция 3: Фильтрация
        main_layout.addWidget(QLabel("<h2>3. Фильтрация данных</h2>"))
        filter_layout = QHBoxLayout()
        filter_layout.addWidget(QLabel("Вариант фильтрации:"))
        self.filter_combo = QComboBox()
        self.filter_combo.addItems([
            "Без фильтрации",
            "Удалить строки с пустым ПЕРВЫМ столбцом",
            "Удалить ПОЛНОСТЬЮ пустые строки",
        ])
        index = self.filter_combo.findText(self.filter_option)
        if index != -1:
            self.filter_combo.setCurrentIndex(index)
        self.filter_combo.currentIndexChanged.connect(self._save_current_settings_to_temp_preset)
        filter_layout.addWidget(self.filter_combo)
        main_layout.addLayout(filter_layout)

        # Секция 4: Сохранение результата
        main_layout.addWidget(QLabel("<h2>4. Сохранение результата</h2>"))
        output_file_layout = QHBoxLayout()
        self.output_file_line_edit = QLineEdit(os.path.join(self.output_directory, self.output_filename))
        self.output_file_line_edit.textChanged.connect(self._save_current_settings_to_temp_preset)
        self.output_file_line_edit.setReadOnly(True)
        self.btn_select_output_file_path = QPushButton("Выбрать файл для сохранения")
        self.btn_select_output_file_path.clicked.connect(self._select_output_file_path)
        output_file_layout.addWidget(self.output_file_line_edit)
        output_file_layout.addWidget(self.btn_select_output_file_path)
        main_layout.addLayout(output_file_layout)

        # Секция 5: Запуск и логи
        main_layout.addWidget(QLabel("<h2>5. Запуск и результат</h2>"))
        run_buttons_layout = QHBoxLayout()
        self.btn_run = QPushButton("Запустить обработку")
        self.btn_run.setStyleSheet("font-size: 18px; padding: 10px; background-color: #4CAF50; color: white;")
        self.btn_run.clicked.connect(self._run_processing)
        run_buttons_layout.addWidget(self.btn_run)
        self.btn_open_output_file = QPushButton("Открыть созданный файл")
        self.btn_open_output_file.setStyleSheet("font-size: 16px; padding: 10px; background-color: #1E90FF; color: white;")
        self.btn_open_output_file.setEnabled(False)
        self.btn_open_output_file.clicked.connect(self._open_last_output_file)
        run_buttons_layout.addWidget(self.btn_open_output_file)
        main_layout.addLayout(run_buttons_layout)

        self.log_output = QTextEdit()
        self.log_output.setReadOnly(True)
        self.log_output.setStyleSheet("background-color: #f0f0f0; border: 1px solid #ccc;")
        self.log_output.setFixedHeight(200)

        log_scroll_area = QScrollArea()
        log_scroll_area.setWidgetResizable(True)
        log_scroll_area.setWidget(self.log_output)
        log_scroll_area.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Expanding)
        main_layout.addWidget(log_scroll_area)

        main_layout.addStretch(1)

        self._update_preset_list()
        self._set_processing_controls_enabled(False)
        self.btn_clear_source_files.setEnabled(bool(self.source_file_paths))

    def _log(self, message: str):
        self.log_output.append(message)
        QApplication.processEvents()
        self.log_output.verticalScrollBar().setValue(self.log_output.verticalScrollBar().maximum())

    def _set_processing_controls_enabled(self, enabled: bool, enable_marker_selection: bool = True):
        self.sheet_combo.setEnabled(enabled)
        self.btn_preview_marker.setEnabled(enabled and enable_marker_selection and bool(self.first_file_path))
        self.num_cols_edit.setEnabled(enabled)
        self.insert_header_checkbox.setEnabled(enabled)
        self.filter_combo.setEnabled(enabled)
        self.btn_run.setEnabled(enabled and bool(self.source_file_paths) and bool(self.selected_sheet_name)
                                and (bool(self.marker_text) or (self.marker_row != -1 and self.marker_col != -1)))

        self.use_coords_checkbox.setEnabled(enabled and enable_marker_selection and bool(self.first_file_path))
        self._coords_checkbox_state_changed(self.use_coords_checkbox.isChecked())

        self.output_file_line_edit.setEnabled(True)
        self.btn_select_output_file_path.setEnabled(True)

    def _coords_checkbox_state_changed(self, checked: bool):
        self.use_cell_coordinates_for_marker = checked
        self.marker_text_edit.setVisible(not checked)
        self.marker_coords_label.setVisible(checked)
        self.marker_coords_edit.setVisible(checked)
        self.marker_text_edit.setReadOnly(True)
        self.marker_coords_edit.setReadOnly(True)

        if hasattr(self, "btn_run"):
            self.btn_run.setEnabled(bool(self.source_file_paths) and bool(self.selected_sheet_name)
                                    and (bool(self.marker_text) or (self.marker_row != -1 and self.marker_col != -1)))

        self._save_current_settings_to_temp_preset()

    def _get_current_settings(self) -> dict:
        return {
            SOURCE_FILES_KEY: self.source_file_paths,
            OUTPUT_DIR_KEY: self.output_directory,
            OUTPUT_FILENAME_KEY: self.output_filename,
            LAST_SELECTED_SHEET_KEY: self.selected_sheet_name,
            MARKER_TEXT_KEY: self.marker_text,
            MARKER_ROW_KEY: self.marker_row,
            MARKER_COL_KEY: self.marker_col,
            USE_CELL_COORD_KEY: self.use_cell_coordinates_for_marker,
            NUM_COLUMNS_KEY: self.num_columns,
            INSERT_HEADER_KEY: self.insert_header,
            FILTER_OPTION_KEY: self.filter_option,
            NUMERIC_COLUMNS_KEY: self.numeric_columns,
        }

    def _apply_settings_to_ui(self, settings: dict):
        self._is_applying_preset = True
        try:
            self.source_file_paths = settings.get(SOURCE_FILES_KEY, [])
            self.output_directory = settings.get(OUTPUT_DIR_KEY, str(Path(__file__).parent.resolve()))
            self.output_filename = settings.get(OUTPUT_FILENAME_KEY, "output.xlsx")
            self.selected_sheet_name = settings.get(LAST_SELECTED_SHEET_KEY)
            self.marker_text = settings.get(MARKER_TEXT_KEY, "")
            self.marker_row = int(settings.get(MARKER_ROW_KEY, -1)) if settings.get(MARKER_ROW_KEY, -1) is not None else -1
            self.marker_col = int(settings.get(MARKER_COL_KEY, -1)) if settings.get(MARKER_COL_KEY, -1) is not None else -1
            self.use_cell_coordinates_for_marker = settings.get(USE_CELL_COORD_KEY, False)
            self.num_columns = settings.get(NUM_COLUMNS_KEY, 5)
            self.insert_header = settings.get(INSERT_HEADER_KEY, True)
            self.filter_option = settings.get(FILTER_OPTION_KEY, "Без фильтрации")
            self.numeric_columns = settings.get(NUMERIC_COLUMNS_KEY, [])

            self.use_coords_checkbox.blockSignals(True)
            self.marker_text_edit.setText(self.marker_text)
            self.marker_coords_edit.setText(f"({self.marker_row+1}, {self.marker_col+1})" if self.marker_row != -1 and self.marker_col != -1 else "")
            self.use_coords_checkbox.setChecked(self.use_cell_coordinates_for_marker)
            self.use_coords_checkbox.blockSignals(False)
            self.num_cols_edit.setText(str(self.num_columns))
            self.insert_header_checkbox.setChecked(self.insert_header)

            index = self.filter_combo.findText(self.filter_option)
            if index != -1:
                self.filter_combo.setCurrentIndex(index)
            else:
                self.filter_combo.setCurrentIndex(0)
                self.filter_option = self.filter_combo.currentText()

            self.output_file_line_edit.blockSignals(True)
            self.output_file_line_edit.setText(os.path.join(self.output_directory, self.output_filename))
            self.output_file_line_edit.blockSignals(False)

            self._update_file_list_display()
            self._load_sheets_and_settings()

            self._set_processing_controls_enabled(True)
            self.marker_coords_edit.setText(f"({self.marker_row+1}, {self.marker_col+1})" if self.marker_row != -1 and self.marker_col != -1 else "")
            self._update_numeric_label()
        finally:
            self._is_applying_preset = False

    def _update_numeric_label(self):
        if self.numeric_columns:
            self.numeric_label.setText(f"Числовые столбцы (номера от начала данных): {', '.join(str(i+1) for i in sorted(self.numeric_columns))}")
        else:
            self.numeric_label.setText("Числовые столбцы: не выбраны")

    def _update_preset_list(self):
        self.preset_combo.blockSignals(True)
        self.preset_combo.clear()
        self.preset_combo.addItems(["<Новый пресет>"] + self.preset_manager.get_all_preset_names())

        if self.preset_manager.last_used_preset_name:
            index = self.preset_combo.findText(self.preset_manager.last_used_preset_name)
            if index != -1:
                self.preset_combo.setCurrentIndex(index)
            else:
                self.preset_combo.setCurrentIndex(0)
        else:
            self.preset_combo.setCurrentIndex(0)

        self.preset_combo.blockSignals(False)

    def _preset_selected(self, index: int):
        preset_name = self.preset_combo.currentText()
        if preset_name == "<Новый пресет>":
            self._log("Выбран <Новый пресет>. Сброс настроек.")
            self.source_file_paths = []
            self.output_directory = str(Path(__file__).parent.resolve())
            self.output_filename = "output.xlsx"
            self.selected_sheet_name = None
            self.marker_text = ""
            self.marker_row = -1
            self.marker_col = -1
            self.use_cell_coordinates_for_marker = False
            self.num_columns = 5
            self.insert_header = True
            self.filter_option = "Без фильтрации"
            self.numeric_columns = []
            self.first_file_path = None

            settings_with_defaults = self._get_current_settings()
            self._apply_settings_to_ui(settings_with_defaults)

        else:
            settings = self.preset_manager.load_preset(preset_name)
            if settings:
                self._log(f"Загружен пресет: '{preset_name}'")
                self._apply_settings_to_ui(settings)
            else:
                self._log(f"⚠ Не удалось загрузить пресет: '{preset_name}'")
                self.preset_combo.setCurrentIndex(0)

    def _save_current_preset(self):
        current_preset_name = self.preset_combo.currentText()
        new_preset_name = ""

        self._sync_ui_to_attributes()

        if current_preset_name == "<Новый пресет>":
            preset_name_input, ok = QInputDialog.getText(self, "Сохранить пресет", "Введите имя для нового пресета:")
            if not ok or not preset_name_input.strip():
                self._log("Сохранение пресета отменено.")
                return
            new_preset_name = preset_name_input.strip()
        else:
            settings_to_save = self._get_current_settings()

            if current_preset_name != "<Новый пресет>" and current_preset_name in self.preset_manager.get_all_preset_names():
                loaded_settings = self.preset_manager.load_preset(current_preset_name)
                if loaded_settings == settings_to_save:
                    self._log(f"Пресет '{current_preset_name}' не изменился. Сохранение отменено.")
                    return

            reply = QMessageBox.question(self, "Сохранить пресет",
                                         f"Пресет '{current_preset_name}' уже существует и изменен. Хотите <b>перезаписать</b> его или <b>сохранить как новый</b>?",
                                         QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                                         QMessageBox.Yes)
            if reply == QMessageBox.Yes:
                new_preset_name = current_preset_name
            elif reply == QMessageBox.No:
                preset_name_input, ok = QInputDialog.getText(self, "Сохранить пресет", "Введите новое имя пресета:", text=f"{current_preset_name}_{datetime.datetime.now().strftime('%H%M%S')}")
                if not ok or not preset_name_input.strip():
                    self._log("Сохранение пресета отменено.")
                    return
                new_preset_name = preset_name_input.strip()
            else:
                self._log("Сохранение пресета отменено.")
                return

        if not new_preset_name:
            self._log("Сохранение пресета отменено (не получено имя).")
            return

        settings_to_save = self._get_current_settings()
        if self.preset_manager.save_preset(new_preset_name, settings_to_save):
            self._log(f"✓ Пресет '{new_preset_name}' успешно сохранен.")
            self._update_preset_list()
            index = self.preset_combo.findText(new_preset_name)
            if index != -1:
                self.preset_combo.setCurrentIndex(index)
        else:
            QMessageBox.critical(self, "Ошибка", f"Не удалось сохранить пресет '{new_preset_name}'.")

    def _delete_current_preset(self):
        current_preset_name = self.preset_combo.currentText()
        if current_preset_name == "<Новый пресет>":
            QMessageBox.warning(self, "Удаление пресета", "Нельзя удалить '<Новый пресет>'.")
            return

        reply = QMessageBox.question(self, "Удалить пресет",
                                     f"Вы уверены, что хотите удалить пресет '{current_preset_name}'?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            if self.preset_manager.delete_preset(current_preset_name):
                self._log(f"✓ Пресет '{current_preset_name}' успешно удален.")
                self._update_preset_list()
                self._preset_selected(0)
            else:
                QMessageBox.critical(self, "Ошибка", f"Не удалось удалить пресет '{current_preset_name}'.")

    def _sync_ui_to_attributes(self):
        self.selected_sheet_name = self.sheet_combo.currentText() if self.sheet_combo.count() > 0 else None
        self.marker_text = self.marker_text_edit.text()

        if self.use_cell_coordinates_for_marker:
            coords_text = self.marker_coords_edit.text().strip()
            match = re.match(r"\(\s*(\d+)\s*,\s*(\d+)\s*\)", coords_text)
            if match:
                self.marker_row = int(match.group(1)) - 1
                self.marker_col = int(match.group(2)) - 1
        else:
            self.marker_row = -1
            self.marker_col = -1

        try:
            self.num_columns = int(self.num_cols_edit.text())
        except ValueError:
            self.num_columns = 5
            self.num_cols_edit.setText("5")

        self.insert_header = self.insert_header_checkbox.isChecked()
        self.filter_option = self.filter_combo.currentText()

        output_full_path = self.output_file_line_edit.text()
        self.output_directory = os.path.dirname(output_full_path)
        self.output_filename = os.path.basename(output_full_path)

    def _save_current_settings_to_temp_preset(self):
        if self._is_applying_preset:
            return
        self._sync_ui_to_attributes()

    def _load_last_used_preset(self):
        if self.preset_manager.last_used_preset_name:
            settings = self.preset_manager.load_preset(self.preset_manager.last_used_preset_name)
            if settings:
                self._log(f"Загружен последний использованный пресет: '{self.preset_manager.last_used_preset_name}'")
                self._apply_settings_to_ui(settings)
            else:
                self._log(f"⚠ Последний использованный пресет '{self.preset_manager.last_used_preset_name}' не найден или поврежден. Загружен <Новый пресет>.")
                self._preset_selected(0)
        else:
            self._log("Последний использованный пресет не найден. Загружен <Новый пресет>.")
            self._preset_selected(0)

    def _select_source_files(self):
        dialog = QFileDialog(self)
        dialog.setFileMode(QFileDialog.ExistingFiles)
        dialog.setNameFilter("Excel Files (*.xlsx *.xlsm)")
        dialog.setWindowTitle("Выберите один или несколько Excel файлов")

        current_files = self.source_file_paths

        if not self.source_file_paths:
            if dialog.exec():
                self.source_file_paths = dialog.selectedFiles()
            else:
                return
        else:
            reply = QMessageBox.question(self, "Добавить файлы?",
                                         "Вы уже выбрали файлы. Хотите <b>добавить</b> к ним новые или <b>заменить</b> текущие?",
                                         QMessageBox.Yes | QMessageBox.No | QMessageBox.Cancel,
                                         QMessageBox.Yes)

            if reply == QMessageBox.Yes:
                if dialog.exec():
                    self.source_file_paths.extend(dialog.selectedFiles())
            elif reply == QMessageBox.No:
                if dialog.exec():
                    self.source_file_paths = dialog.selectedFiles()
            else:
                return

        if not self.source_file_paths and current_files:
            self._log("Список исходных файлов был очищен.")

        old_count = len(self.source_file_paths)
        self.source_file_paths = get_excel_files(self.source_file_paths)
        new_count = len(self.source_file_paths)

        if new_count < old_count:
            self._log(f"⚠ Убраны дубликаты и несуществующие/не-Excel пути. Итого: {new_count} файлов.")

        self._update_file_list_display()
        self._load_sheets_and_settings()

    def _clear_source_files(self):
        reply = QMessageBox.question(self, "Очистить список?",
                                     "Вы уверены, что хотите очистить список выбранных файлов?",
                                     QMessageBox.Yes | QMessageBox.No)
        if reply == QMessageBox.Yes:
            self.source_file_paths = []
            self.first_file_path = None

            self.selected_sheet_name = None
            self.marker_text = ""
            self.marker_row = -1
            self.marker_col = -1
            self.use_cell_coordinates_for_marker = False
            self.numeric_columns = []

            self.sheet_combo.clear()
            self.marker_text_edit.clear()
            self.marker_coords_edit.clear()
            self.use_coords_checkbox.setChecked(False)
            self._update_numeric_label()

            self._update_file_list_display()
            self._load_sheets_and_settings()
            self._set_processing_controls_enabled(False)
            self._log("Список исходных файлов очищен.")
            self.btn_clear_source_files.setEnabled(False)
            self.btn_open_output_file.setEnabled(False)

    def _select_output_file_path(self):
        default_path = os.path.join(self.output_directory, self.output_filename)
        file_path, _ = QFileDialog.getSaveFileName(self, "Сохранить результат как",
                                                   default_path,
                                                   "Excel Files (*.xlsx)")
        if file_path:
            self.output_file_line_edit.setText(file_path)
            self._log(f"✓ Путь и имя для сохранения изменены на: {file_path}")

    def _update_file_list_display(self):
        self.btn_open_output_file.setEnabled(False)

        if not self.source_file_paths:
            self.file_list_text.setText("❌ Файлы для обработки не выбраны.")
            self._set_processing_controls_enabled(False)
            self.btn_clear_source_files.setEnabled(False)
            self.first_file_path = None
            return

        file_names_display = "\n".join([f"{i+1}. {os.path.basename(f)}" for i, f in enumerate(self.source_file_paths)])
        self.file_list_text.setText(f"✓ Выбрано файлов: {len(self.source_file_paths)}\n{file_names_display}")
        self._log(f"✓ Выбрано {len(self.source_file_paths)} файлов.")
        self.btn_clear_source_files.setEnabled(True)

        self.first_file_path = self.source_file_paths[0]
        self._set_processing_controls_enabled(True)

    def _load_sheets_and_settings(self):
        self.sheet_combo.blockSignals(True)
        self.sheet_combo.clear()

        desired_sheet_name = self.selected_sheet_name

        if not self.first_file_path or not os.path.exists(self.first_file_path):
            self.sheet_combo.addItem("Нет файла для загрузки листов")
            self._set_processing_controls_enabled(False, enable_marker_selection=False)
            self._log("⚠ Нет первого файла или файл не найден. Невозможно загрузить листы.")
            self.selected_sheet_name = None
            self.sheet_combo.blockSignals(False)
            self._save_current_settings_to_temp_preset()
            return

        try:
            xls = pd.ExcelFile(self.first_file_path)
            sheets_from_file = xls.sheet_names
            if not sheets_from_file:
                self.sheet_combo.addItem("Нет листов")
                self._set_processing_controls_enabled(False, enable_marker_selection=False)
                self._log(f"❌ В файле '{os.path.basename(self.first_file_path)}' нет листов.")
                self.selected_sheet_name = None
                self.sheet_combo.blockSignals(False)
                self._save_current_settings_to_temp_preset()
                return

            self.sheet_combo.addItems(sheets_from_file)
            self._log(f"✓ Загружены листы из '{os.path.basename(self.first_file_path)}'.")

            if desired_sheet_name and desired_sheet_name in sheets_from_file:
                index = self.sheet_combo.findText(desired_sheet_name)
                if index != -1:
                    self.sheet_combo.setCurrentIndex(index)
                else:
                    self.sheet_combo.setCurrentIndex(0)
            else:
                self.sheet_combo.setCurrentIndex(0)

            self.selected_sheet_name = self.sheet_combo.currentText()

            self.marker_text_edit.setText(self.marker_text)
            self.marker_coords_edit.setText(f"({self.marker_row+1}, {self.marker_col+1})" if self.marker_row != -1 and self.marker_col != -1 else "")
            self.use_coords_checkbox.blockSignals(True)
            self.use_coords_checkbox.setChecked(self.use_cell_coordinates_for_marker)
            self.use_coords_checkbox.blockSignals(False)
            self.num_cols_edit.setText(str(self.num_columns))
            self.insert_header_checkbox.setChecked(self.insert_header)
            index = self.filter_combo.findText(self.filter_option)
            if index != -1:
                self.filter_combo.setCurrentIndex(index)
            else:
                self.filter_combo.setCurrentIndex(0)
                self.filter_option = self.filter_combo.currentText()

            self.output_file_line_edit.blockSignals(True)
            self.output_file_line_edit.setText(os.path.join(self.output_directory, self.output_filename))
            self.output_file_line_edit.blockSignals(False)

            self._set_processing_controls_enabled(True)

        except Exception as e:
            self.sheet_combo.addItem("Ошибка загрузки листов")
            self._set_processing_controls_enabled(False, enable_marker_selection=False)
            self._log(f"❌ Ошибка при загрузке листов из '{os.path.basename(self.first_file_path)}': {e}")
            sys.excepthook(type(e), e, e.__traceback__)
        finally:
            self.sheet_combo.blockSignals(False)
            self._save_current_settings_to_temp_preset()

    def _sheet_selected(self, index: int):
        self._sync_ui_to_attributes()
        self._log(f"Выбран лист: {self.selected_sheet_name}")
        self._set_processing_controls_enabled(True)

    def _open_preview_dialog(self):
        if not self.first_file_path or not os.path.exists(self.first_file_path) or not self.selected_sheet_name:
            QMessageBox.warning(self, "Предупреждение", "Сначала выберите хотя бы один Excel файл и лист Excel.")
            return

        marker_text = self.marker_text if self.marker_text is not None else ""

        dialog = PreviewDialog(
            self.first_file_path,
            self.selected_sheet_name,
            self.marker_row,
            self.marker_col,
            self.numeric_columns,
            marker_text,
            self
        )
        if dialog.exec() == QDialog.Accepted:
            if self.use_cell_coordinates_for_marker:
                self.marker_row = dialog.selected_row
                self.marker_col = dialog.selected_col
                self.marker_text = dialog.selected_cell_text
                self._log(f"✓ Маркер выбран по координатам: строка {self.marker_row + 1}, столбец {self.marker_col + 1} (текст: '{self.marker_text}').")
            else:
                self.marker_text = dialog.selected_cell_text
                self.marker_row = -1
                self.marker_col = -1
                self._log(f"✓ Маркер '{self.marker_text}' выбран для поиска.")

            self.numeric_columns = dialog.numeric_columns_rel
            self._update_numeric_label()
            self._log(f"✓ Выбраны числовые столбцы (номера от начала данных): {', '.join(str(i+1) for i in sorted(self.numeric_columns)) if self.numeric_columns else 'нет'}")

            self.marker_text_edit.setText(self.marker_text)
            self.marker_coords_edit.setText(f"({self.marker_row+1}, {self.marker_col+1})" if (self.marker_row != -1 and self.marker_col != -1) else "")
            self._save_current_settings_to_temp_preset()
            self._set_processing_controls_enabled(True)
        else:
            self._log("Отменено: выбор маркера/координат не произведен.")

    def _open_last_output_file(self):
        if self.last_output_file_path and os.path.exists(self.last_output_file_path):
            try:
                if sys.platform == "win32":
                    os.startfile(self.last_output_file_path)
                elif sys.platform == "darwin":
                    subprocess.call(["open", self.last_output_file_path])
                else:
                    subprocess.call(["xdg-open", self.last_output_file_path])
                self._log(f"✓ Открыт файл: {self.last_output_file_path}")
            except Exception as e:
                self._log(f"❌ Не удалось открыть файл '{self.last_output_file_path}': {e}")
                QMessageBox.critical(self, "Ошибка", f"Не удалось открыть файл: {e}")
        else:
            QMessageBox.warning(self, "Ошибка", "Файл для открытия не найден или еще не создан.")

    def _run_processing(self):
        self.log_output.clear()
        self.btn_open_output_file.setEnabled(False)
        self.last_output_file_path = None

        self._log("=" * 60)
        self._log("НАЧАЛО ОБРАБОТКИ ФАЙЛОВ")
        self._log("=" * 60)

        self._sync_ui_to_attributes()

        if not self.source_file_paths:
            self._log("❌ Нет выбранных файлов для обработки.")
            QMessageBox.critical(self, "Ошибка", "Нет файлов для обработки.")
            return

        selected_sheet = self.selected_sheet_name
        if not selected_sheet:
            self._log("❌ Не выбран лист для обработки.")
            QMessageBox.critical(self, "Ошибка", "Выберите лист для обработки.")
            return

        num_columns = self.num_columns
        if num_columns <= 0:
            self._log("❌ Количество столбцов должно быть больше 0.")
            QMessageBox.critical(self, "Ошибка", "Количество столбцов должно быть больше 0.")
            return

        if self.use_cell_coordinates_for_marker:
            if self.marker_row == -1 or self.marker_col == -1:
                self._log("❌ Активирован режим использования координат, но координаты не указаны.")
                QMessageBox.critical(self, "Ошибка", "Укажите координаты ячейки.")
                return
        else:
            if not self.marker_text:
                self._log("❌ Не указан текст маркера начала данных.")
                QMessageBox.critical(self, "Ошибка", "Укажите текст маркера начала данных.")
                return

        insert_header = self.insert_header
        filter_option = self.filter_option

        output_path = os.path.join(self.output_directory, self.output_filename)

        if not output_path:
            self._log("❌ Не указан путь для сохранения файла.")
            QMessageBox.critical(self, "Ошибка", "Укажите путь для сохранения файла.")
            return

        output_dir = self.output_directory

        if not os.path.isdir(output_dir):
            reply = QMessageBox.question(self, "Создать директорию?",
                                         f"Директория '{output_dir}' не существует. Создать её?",
                                         QMessageBox.Yes | QMessageBox.No)
            if reply == QMessageBox.Yes:
                try:
                    os.makedirs(output_dir)
                    self._log(f"✓ Директория '{output_dir}' успешно создана.")
                except Exception as e:
                    self._log(f"❌ Не удалось создать директорию '{output_dir}': {e}")
                    QMessageBox.critical(self, "Ошибка", f"Не удалось создать директорию для сохранения: {e}")
                    return
            else:
                self._log("❌ Пользователь отменил создание директории.")
                return

        self._log("\n----- Настройки обработки -----")
        self._log(f"✓ Путь к выходному файлу: {output_path}")
        self._log(f"✓ Лист: {selected_sheet}")
        if self.use_cell_coordinates_for_marker:
            self._log(f"✓ Маркер: по координатам (строка {self.marker_row+1}, столбец {self.marker_col+1})")
        else:
            self._log(f"✓ Маркер: по тексту '{self.marker_text}'")
        self._log(f"✓ Столбцов: {num_columns}")
        self._log(f"✓ Вставлять заголовок: {'Да' if insert_header else 'Нет'}")
        self._log(f"✓ Фильтрация: {filter_option}")
        self._log(f"✓ Числовые столбцы (номера от начала данных): {', '.join(str(i+1) for i in sorted(self.numeric_columns)) if self.numeric_columns else 'нет'}")
        self._log("-------------------------------")
        self._log("\n<b>Список обрабатываемых файлов:</b>")
        for i, f_path in enumerate(self.source_file_paths):
            self._log(f"  {i+1}. {os.path.basename(f_path)}")
        self._log("-----------------------------\n")

        all_dataframes = []
        header_row_values = None

        for i, file_path in enumerate(self.source_file_paths, 1):
            if not os.path.exists(file_path):
                self._log(f"[ШАГ] ❌ Файл не найден, пропущен: {os.path.basename(file_path)}")
                continue

            self._log(f"\n[ШАГ] Обработка файла {i}/{len(self.source_file_paths)}: {os.path.basename(file_path)}")

            df_data, header_row_values = extract_data_from_file(
                file_path=file_path,
                sheet_name=selected_sheet,
                marker_text=self.marker_text,
                marker_row=self.marker_row,
                marker_col=self.marker_col,
                use_cell_coordinates_for_marker=self.use_cell_coordinates_for_marker,
                num_columns=num_columns,
                insert_header=insert_header,
                header_row_values= header_row_values,
                numeric_columns=self.numeric_columns,   # относительные индексы
                log_func=self._log
            )

            if df_data is not None:
                all_dataframes.append(df_data)
            QApplication.processEvents()

        if not all_dataframes:
            self._log("\n❌ Нет данных для объединения. Программа завершена.")
            QMessageBox.warning(self, "Нет данных", "Не удалось извлечь данные из файлов.")
            return

        self._log("\n[ШАГ] Объединение данных...")
        final_df = pd.concat(all_dataframes, ignore_index=True)
        self._log(f"✓ Объединено: {len(final_df)} строк, {len(final_df.columns)} столбцов до фильтрации.")

        header_row_for_filter_index = 0

        if insert_header and header_row_values:
        # if  header_row_values:
            header_df = pd.DataFrame([header_row_values], columns=final_df.columns[:len(header_row_values)])

            if "Source_File" in final_df.columns and "Source_File" not in header_df.columns:
                if len(header_df.columns) < len(final_df.columns):
                    header_df["Source_File"] = "Source_File"

            final_df = pd.concat([header_df, final_df], ignore_index=True)
            self._log("✓ Пользовательский заголовок успешно вставлен как первая строка.")
            header_row_for_filter_index = 1
        elif insert_header and not header_row_values:
            self._log("❌ Заголовок НЕ был вставлен, так как не удалось его извлечь.")

        self._log("\n[ШАГ] Применение фильтрации...")
   
        final_df = filter_dataframe(final_df, filter_option, log_func=self._log)
        self._log(f"✓ Итого строк после фильтрации: {len(final_df) - (1 if insert_header and header_row_values else 0)}")
        self._log(f"✓ Итого строк в DataFrame для сохранения: {len(final_df)}")

        self._log("\n[ШАГ] Сохранение результата...")
        output_main_sheet_name = OUTPUT_SHEET_NAME

        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                final_df.to_excel(writer, sheet_name=output_main_sheet_name, index=False, header=not (insert_header and header_row_values))

            self._log(f"✓ Применение автофильтра к листу '{output_main_sheet_name}'...")
            if header_row_for_filter_index == 0:
                header_row_for_filter_index = 1

            apply_auto_filter(output_path, output_main_sheet_name, header_row_for_filter_index, log_func=self._log)

            self.last_output_file_path = output_path
            self.btn_open_output_file.setEnabled(True)

            self._log(f"✓ Данные сохранены: {output_path}")
            self._log(f"✓ Итого строк в выходном файле (с учетом заголовков и фильтрации): {len(final_df)}")
            self._log(f"✓ Итого столбцов в выходном файле: {len(final_df.columns)}")
            self._log("\n" + "=" * 60)
            self._log("✅ ПРОГРАММА УСПЕШНО ЗАВЕРШЕНА!")
            self._log("Нажми кнопку 'Открыть созданный файл' для просмотра.")
            self._log("=" * 60)
            QMessageBox.information(self, "Завершено", f"Обработка завершена. Результат сохранен в:\n{output_path}")
        except Exception as e:
            self._log(f"❌ Ошибка при сохранении файла или применении автофильтра: {e}")
            QMessageBox.critical(self, "Ошибка сохранения", f"Не удалось сохранить файл или применить автофильтр: {e}")
            sys.excepthook(type(e), e, e.__traceback__)

if __name__ == "__main__":
    def handle_exception(exc_type, exc_value, exc_traceback):
        if issubclass(exc_type, KeyboardInterrupt):
            sys.__excepthook__(exc_type, exc_value, exc_traceback)
            return
        QMessageBox.critical(None, "Необработанное исключение",
                             f"Произошла непредвиденная ошибка:\n{exc_value}",
                             QMessageBox.Ok)
        sys.__excepthook__(exc_type, exc_value, exc_traceback)

    sys.excepthook = handle_exception

    app = QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())
