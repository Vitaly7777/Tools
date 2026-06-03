import os
import platform
from gui.ui_qt5 import login_ui
from PyQt5.QtWidgets import QMessageBox
if platform.system() == "Linux":
    import smbclient
    from smbprotocol.exceptions import SMBResponseException

import pandas as pd
import logging
from openpyxl import load_workbook

class MyCustomError(Exception):
    def __init__(self, message="Произошла пользовательская ошибка"):
        self.message = message
        super().__init__(self.message)

def load_smb_credentials(path:str = "\\\\pr.rt.ru\\FSPR"):
    '''
    Функция производит настройку smbclient,
    проверяет правильность вводалогина пароля
    Возвращает кортеж (логин, пароль) или None
    '''
    system = platform.system()
    if system == "Linux":

        while True:
            secret = login_ui("Введите доменный логин/пароль")
            if secret:
                username, password = secret   
                smbclient.ClientConfig(username=username, password=password)

                try:
                    smbclient.listdir(path)
                    username, password = None, None
                    return secret
        
                except SMBResponseException:
                    QMessageBox.critical(
                        None,
                        "Ошибка входа",
                        "Неверный логин или пароль. Повторите попытку.",
                    )
                    continue

            else:
                return None



def read_df(file_path: str, **options: dict) -> pd.DataFrame:
    '''
    Функция чтения Excel-файла с сетевого(Samba) ресурса.
    Возвращает датафрейм или генерирует пользоваельскую ошибку или запись ошибки в лог и None
    '''
    # logger = logging.getLogger('read_df')
    # logger.debug('Начало выполнения подпрограммы read_df.')
    system = platform.system()
    if system == "Linux":
        try:
            with smbclient.open_file(file_path, mode="rb") as file:
                return pd.read_excel(file, **options)
        except Exception as e:
            # QMessageBox.warning(
            #         macro.parent, "Ошибка", f'Ошибка чтения!\n{e}')
            raise MyCustomError ('Ошибка чтения!\n{e}')
            # logger.error(f'Ошибка чтения!\n{e}')
            return None
    else:
        try:
            df = pd.read_excel(file_path, **options)
            return df
        except Exception as e:
            # QMessageBox.warning(
            #         macro.parent, "Ошибка", f'Ошибка чтения!\n{e}')
            raise MyCustomError ('Ошибка чтения!\n{e}')
            logger.error(f'Ошибка чтения!\n{e}')
            return None

def write_df(df:pd.DataFrame, output_file_path: str, sheet_name: str):
    # logger = logging.getLogger('write_df')
    # logger.debug('Начало выполнения подпрограммы write_df.')
    system = platform.system()
    try:
        if system == "Linux":
            with smbclient.open_file(output_file_path, mode="wb") as file:
                df.to_excel(file, index=False, sheet_name=sheet_name) 

        else:
            df.to_excel(output_file_path, index=False, sheet_name=sheet_name)  
    except Exception as e:
        raise MyCustomError ('Ошибка записи!\n{e}')
        # logger.error(f'Ошибка записи!\n{e}')

 
def read_named_range(file_path, range_name):
    '''
    Функция для чтения именованного диапазона из Excel-файла и загрузки в DataFrame.

    Параметры:
    - file_path: str, путь к файлу (.xlsx или  .xlsm ).
    - range_name: str, имя именованного диапазона (например, 'MyRange').

    Возвращает:
    - DataFrame с данными из диапазона.
    - Если ошибка, возвращает None и выводит сообщение.

    Исключения:
    - Проверяет расширение файла.
    - Обрабатывает ошибки, если файл не найден или диапазон не существует.
    '''
   
    # Шаг 1: Проверить расширение файла
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in ['.xlsx', '.xlsm']:
        raise ValueError('Файл должен иметь расширение .xlsx  или .xlsm')
    
    # Настройка логера (если нужно)
    # logger = logging.getLogger('read_named_range')
    # logger.debug('Начало выполнения подпрограммы read_named_range')
    try:
        # Шаг 2: Загрузить workbook с помощью openpyxl
        system = platform.system()
        if system == 'Linux': 
            with smbclient.open_file(file_path, mode='rb') as file:
                wb = load_workbook(file, data_only=True)
        else:
            wb = load_workbook(file_path, read_only=True, data_only=True)  # data_only=True для получения значений, а не формул
        # Шаг 3: Получить именованный диапазон
        if range_name in wb.defined_names:
            defined_name = wb.defined_names[range_name]
            destinations = list(defined_name.destinations)
            # Получите лист и диапазон ячеек
            sheet_name = destinations[0][0]  # Лист, на котором находится диапазон
            cell_range = destinations[0][1]  # Диапазон ячеек (например, 'A1:C3')
            # logger.debug(f'sheet_name ={sheet_name}, cell_range={cell_range}')
            # Получите данные из диапазона
            sheet = wb[sheet_name]
            data = []
            for row in sheet[cell_range]:
                data.append([cell.value for cell in row])
            # Первая строка — это заголовки столбцов
            # headers = list(data[0])
            headers = data[0]
            # logger.debug(f'headers={headers}')
            rows = data[1:]
            # Преобразуем данные в DataFrame
            df = pd.DataFrame(rows, columns=headers)    
            return df  # Возвращаем DataFrame
        else:
            raise ValueError(f"Именованный диапазон '{range_name}' не найден в файле.")
    except FileNotFoundError:
        # print(f"Ошибка: Файл '{file_path}' не найден.")
        # logger.error(f"Ошибка: Файл '{file_path}' не найден.")
        return None
    except Exception as e:
        # print(f'Общая ошибка: {e}')
        # logger.error(f'Общая ошибка: {e}')
        return None

def read_list_obj(file_path:str, list_object_name:str):
    '''
    Функция для чтения умной таблицы из Excel-файла и загрузки в DataFrame.

    Параметры:
    - file_path: str, путь к файлу (.xlsx или  .xlsm ).
    - list_object_name: str, имя именованного диапазона (например, 'ListObj02').

    Возвращает:
    - DataFrame с данными из диапазона.
    - Если ошибка, возвращает None и выводит сообщение.

    Исключения:
    - Проверяет расширение файла.
    - Обрабатывает ошибки, если файл не найден или диапазон не существует.
    '''
    # from openpyxl import load_workbook
    # Шаг 1: Проверить расширение файла
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in ['.xlsx', '.xlsm']:
        raise ValueError('Файл должен иметь расширение .xlsx  или .xlsm')
    
    # Настройка логера (если нужно)
    # logger = logging.getLogger('read_list_obj')
    # logger.debug('Начало выполнения подпрограммы read_list_obj')
    try:
        # Шаг 2: Загрузить workbook с помощью openpyxl
        system = platform.system()
        if system == 'Linux': 
            with smbclient.open_file(file_path, mode='rb') as file:
                wb = load_workbook(file, data_only=True)
        else:
            wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            # logger.debug(f'sheet_name = {sheet_name}')
            sheet = wb[sheet_name]
            # Проверяем, есть ли в листе таблицы
            if sheet.tables:
                # print(f"Лист '{sheet_name}' содержит 'умную таблицу'.")
                # logger.debug(f'sheet_name = {sheet_name} содержит \n{sheet.tables}')
                # Выводим информацию о таблицах
                for table_name, addr in sheet.tables.items():
                    if  table_name == list_object_name:
                        # Получите диапазон таблицы
                        # Извлеките данные из диапазона
                        data = []
                        for row in sheet[addr]:
                            data.append([cell.value for cell in row])
                        # Первая строка — это заголовки столбцов
                        headers = data[0]
                        # Остальные строки — это данные
                        rows = data[1:]
                        # Создаем DataFrame
                        df = pd.DataFrame(rows, columns=headers)
                        return df
        
        raise ValueError(f"Умная таблица '{list_object_name}' не найдена в файле:\n{file_path}.")
    
    except FileNotFoundError:
        # print(f"Ошибка: Файл '{file_path}' не найден.")
        # logger.error(f"Ошибка: Файл '{file_path}' не найден.")
        return None
    except Exception as e:
        # print(f'Общая ошибка: {e}')
        # logger.error(f'Общая ошибка: {e}')
        return None               
