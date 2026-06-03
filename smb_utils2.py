import os
import sys
# from typing import Tuple
from pathlib import Path
import platform
from gui.ui_qt5 import login_ui, ensure_qapp
from PyQt5.QtWidgets import QMessageBox

import pandas as pd
import logging
from openpyxl import load_workbook

SHARE_PATH = '//pr.rt.ru/fspr/руз/'
MNT_PATH = '/mnt/dfs/РУЗ/'

def test_mount(path_mnt:str = '/mnt/dfs/РУЗ') -> bool:
    return Path(path_mnt).is_dir()

# if platform.system() == "Linux" and not test_mount(path_mnt=MNT_PATH):   
#     import smbclient
#     from smbprotocol.exceptions import SMBResponseException

class MyCustomError(Exception):
    def __init__(self, message="Произошла пользовательская ошибка"):
        self.message = message
        super().__init__(self.message)



def replaced_path_mnt(path:str, share_path:str = SHARE_PATH, mnt_path:str = MNT_PATH):
    if isinstance(path, str):
        return path.replace('\\', '/').lower().replace(share_path.lower(), mnt_path)
    else:
        return path        

def load_smb_credentials(path:str = "\\\\pr.rt.ru\\FSPR"):
    '''
    Функция производит настройку smbclient,
    проверяет правильность вводалогина пароля
    Возвращает кортеж (логин, пароль) или None
    '''
    import smbclient
    from smbprotocol.exceptions import SMBResponseException
    while True:
        secret = login_ui("Введите доменный логин/пароль")
        if secret:
            username, password = secret   
            smbclient.ClientConfig(username=username, password=password)

            try:
                smbclient.listdir(path)
                username, password = None, None
                return secret
    
            except SMBResponseException:
                QMessageBox.critical(
                    None,
                    "Ошибка входа",
                    "Неверный логин или пароль. Повторите попытку.",
                )
                continue
        else:
            return None

def mnt_or_smb(mnt_path:str = MNT_PATH):
    if sys.platform.startswith("linux") and not test_mount(mnt_path):
        sec = load_smb_credentials() 
    elif sys.platform.startswith("linux") and test_mount(mnt_path):
        sec = False
    else:
        sec = True
    return sec

def smb_read_df(file_path: str, **options: dict) -> pd.DataFrame:
    '''
    Функция чтения Excel-файла с сетевого(Samba) ресурса.
    Возвращает датафрейм или генерирует пользоваельскую ошибку или запись ошибки в лог и None
    '''
    import smbclient
    try:
        with smbclient.open_file(file_path, mode="rb") as file:
            return pd.read_excel(file, **options)
    except Exception as e:
        processing_exception(f'Ошибка чтения!', e)
            # # QMessageBox.warning(

            
def processing_exception(title:str, exc:str):
    # QMessageBox.warning(
    # macro.parent, "Ошибка", f'Ошибка чтения!\n{e}')  
    # logger.error(f'Ошибка чтения!\n{e}')
    raise MyCustomError (f'Ошибка чтения!\n{exc}')
    return None

def mnt_read_df(file_path: str, **options: dict) -> pd.DataFrame:
    try:
        df = pd.read_excel(file_path, **options)
        return df
    except Exception as e:
        # raise e
        processing_exception(f'Ошибка чтения!', e)

def load_df(sec, path, **options:dict):

    if isinstance(sec, tuple):
        try:
            my_df = smb_read_df(path, **options)
            return my_df
        except Exception as e:
            QMessageBox.critical(
                        None,
                        "Ошибка!",
                        f"{e}",
                    )
    elif isinstance(sec, bool):
        try:
            my_df = mnt_read_df(path, **options)
            return my_df
        except Exception as e:
            QMessageBox.critical(
                        None,
                        "Ошибка!",
                        f"{e}",
                    )
                
    else:
        QMessageBox.critical(
                        None,
                        "Заверпшение работы",
                        "Пользователь не ввел логин/пароль. Заверпшение работы.",
                    )

def smb_write_df(df:pd.DataFrame, file_path:str, **options: dict) -> pd.DataFrame:
    '''
    Функция записи Excel-файла на сетевой(Samba) ресурс.
    Возвращает при ошибки генерирует пользоваельскую ошибку или запись ошибки в лог и None
    '''
    import smbclient
    try:
        with smbclient.open_file(file_path, mode="wb") as file:
            df.to_excel(excel_writer=file, **options) 
    except Exception as e:
        processing_exception(f'Ошибка записи SAMBA!', e)
            # # QMessageBox.warning( )
            # 
def mnt_write_df(df:pd.DataFrame, output_file_path: str, **options: dict):
    try:
        df.to_excel(excel_writer=output_file_path, index=False, **options )
    except Exception as e:
        # raise e
        processing_exception(f'Ошибка записи!', e) 

def write_df(sec, df:pd.DataFrame, output_file_path: str, **options):
    # logger = logging.getLogger('write_df')
    # logger.debug('Начало выполнения подпрограммы write_df.')
    if isinstance(sec, tuple):
        try:
            smb_write_df(df, output_file_path, **options)
        except Exception as e:
            QMessageBox.critical(
                        None,
                        "Ошибка!",
                        f"{e}",
                    )
    elif isinstance(sec, bool):
        try:
           mnt_read_df(df, output_file_path, **options)

        except Exception as e:
            QMessageBox.critical(
                        None,
                        "Ошибка!",
                        f"{e}",
                    )
                
    else:
        QMessageBox.critical(
                        None,
                        "Заверпшение работы",
                        "Пользователь не ввел логин/пароль. Заверпшение работы.",
                    )


# def write_df(df:pd.DataFrame, output_file_path: str, sheet_name: str):
#     # logger = logging.getLogger('write_df')
#     # logger.debug('Начало выполнения подпрограммы write_df.')
#     system = platform.system()
#     try:
#         if system == "Linux":
#             with smbclient.open_file(output_file_path, mode="wb") as file:
#                 df.to_excel(file, index=False, sheet_name=sheet_name) 

#         else:
#             df.to_excel(output_file_path, index=False, sheet_name=sheet_name)  
#     except Exception as e:
#         raise MyCustomError ('Ошибка записи!\n{e}')
#         # logger.error(f'Ошибка записи!\n{e}')

 
def read_named_range(file_path, range_name):
    '''
    Функция для чтения именованного диапазона из Excel-файла и загрузки в DataFrame.

    Параметры:
    - file_path: str, путь к файлу (.xlsx или  .xlsm ).
    - range_name: str, имя именованного диапазона (например, 'MyRange').

    Возвращает:
    - DataFrame с данными из диапазона.
    - Если ошибка, возвращает None и выводит сообщение.

    Исключения:
    - Проверяет расширение файла.
    - Обрабатывает ошибки, если файл не найден или диапазон не существует.
    '''
   
    # Шаг 1: Проверить расширение файла
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in ['.xlsx', '.xlsm']:
        raise ValueError('Файл должен иметь расширение .xlsx  или .xlsm')
    
    # Настройка логера (если нужно)
    # logger = logging.getLogger('read_named_range')
    # logger.debug('Начало выполнения подпрограммы read_named_range')
    try:
        # Шаг 2: Загрузить workbook с помощью openpyxl
        system = platform.system()
        if system == 'Linux': 
            with smbclient.open_file(file_path, mode='rb') as file:
                wb = load_workbook(file, data_only=True)
        else:
            wb = load_workbook(file_path, read_only=True, data_only=True)  # data_only=True для получения значений, а не формул
        # Шаг 3: Получить именованный диапазон
        if range_name in wb.defined_names:
            defined_name = wb.defined_names[range_name]
            destinations = list(defined_name.destinations)
            # Получите лист и диапазон ячеек
            sheet_name = destinations[0][0]  # Лист, на котором находится диапазон
            cell_range = destinations[0][1]  # Диапазон ячеек (например, 'A1:C3')
            # logger.debug(f'sheet_name ={sheet_name}, cell_range={cell_range}')
            # Получите данные из диапазона
            sheet = wb[sheet_name]
            data = []
            for row in sheet[cell_range]:
                data.append([cell.value for cell in row])
            # Первая строка — это заголовки столбцов
            # headers = list(data[0])
            headers = data[0]
            # logger.debug(f'headers={headers}')
            rows = data[1:]
            # Преобразуем данные в DataFrame
            df = pd.DataFrame(rows, columns=headers)    
            return df  # Возвращаем DataFrame
        else:
            raise ValueError(f"Именованный диапазон '{range_name}' не найден в файле.")
    except FileNotFoundError:
        # print(f"Ошибка: Файл '{file_path}' не найден.")
        # logger.error(f"Ошибка: Файл '{file_path}' не найден.")
        return None
    except Exception as e:
        # print(f'Общая ошибка: {e}')
        # logger.error(f'Общая ошибка: {e}')
        return None

def read_list_obj(file_path:str, list_object_name:str):
    '''
    Функция для чтения умной таблицы из Excel-файла и загрузки в DataFrame.

    Параметры:
    - file_path: str, путь к файлу (.xlsx или  .xlsm ).
    - list_object_name: str, имя именованного диапазона (например, 'ListObj02').

    Возвращает:
    - DataFrame с данными из диапазона.
    - Если ошибка, возвращает None и выводит сообщение.

    Исключения:
    - Проверяет расширение файла.
    - Обрабатывает ошибки, если файл не найден или диапазон не существует.
    '''
    # from openpyxl import load_workbook
    # Шаг 1: Проверить расширение файла
    extension = os.path.splitext(file_path)[1].lower()
    if extension not in ['.xlsx', '.xlsm']:
        raise ValueError('Файл должен иметь расширение .xlsx  или .xlsm')
    
    # Настройка логера (если нужно)
    # logger = logging.getLogger('read_list_obj')
    # logger.debug('Начало выполнения подпрограммы read_list_obj')
    try:
        # Шаг 2: Загрузить workbook с помощью openpyxl
        system = platform.system()
        if system == 'Linux': 
            with smbclient.open_file(file_path, mode='rb') as file:
                wb = load_workbook(file, data_only=True)
        else:
            wb = load_workbook(file_path, data_only=True)

        for sheet_name in wb.sheetnames:
            # logger.debug(f'sheet_name = {sheet_name}')
            sheet = wb[sheet_name]
            # Проверяем, есть ли в листе таблицы
            if sheet.tables:
                # print(f"Лист '{sheet_name}' содержит 'умную таблицу'.")
                # logger.debug(f'sheet_name = {sheet_name} содержит \n{sheet.tables}')
                # Выводим информацию о таблицах
                for table_name, addr in sheet.tables.items():
                    if  table_name == list_object_name:
                        # Получите диапазон таблицы
                        # Извлеките данные из диапазона
                        data = []
                        for row in sheet[addr]:
                            data.append([cell.value for cell in row])
                        # Первая строка — это заголовки столбцов
                        headers = data[0]
                        # Остальные строки — это данные
                        rows = data[1:]
                        # Создаем DataFrame
                        df = pd.DataFrame(rows, columns=headers)
                        return df
        
        raise ValueError(f"Умная таблица '{list_object_name}' не найдена в файле:\n{file_path}.")
    
    except FileNotFoundError:
        # print(f"Ошибка: Файл '{file_path}' не найден.")
        # logger.error(f"Ошибка: Файл '{file_path}' не найден.")
        return None
    except Exception as e:
        # print(f'Общая ошибка: {e}')
        # logger.error(f'Общая ошибка: {e}')
        return None               
