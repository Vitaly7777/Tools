
import os
import sys
import pandas as pd


import argparse
import json
import logging
from collections import Counter
from concurrent.futures import ThreadPoolExecutor, as_completed
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional
import ezodf


class CutFilesProcessor:
    """Основной класс обработки файлов."""

    def __init__(self, config_path: Optional[Path] = None):
        """
        Инициализация процессора.

        Args:
            config_path: Путь к JSON-файлу конфигурации.
        """
        # self.work_dir = Path(__file__).resolve().parent
        if getattr(sys, 'frozen', False):
            self.work_dir  = Path(sys.executable).parent
        else:
            self.work_dir  = Path(__file__).resolve().parent.parent 

        self.logger = logging.getLogger(__name__)
        if config_path is None:
            default_config_path = self.work_dir / "config.json"
            if default_config_path.exists():
                config_path = default_config_path

        self.config = self._load_config(config_path)
        self.stats: List[Dict[str, Any]] = []

    def _load_config(self, config_path: Optional[Path]) -> Dict[str, Any]:
        """
        Загрузка конфигурации из JSON и аргументов командной строки.
        Приоритет: аргументы CLI > JSON > значения по умолчанию.
        """
        default_config = {
            "source_folder": "data",
            "output_folder": "",
            "rows_to_keep": 50,
            "preserve_formatting": False,
            "max_workers": 1,
            "log_level": "INFO",
            "log_file": "cut_files.log",
            "stats_file": "processing_stats.xlsx"
        }

        json_config = {}

        if config_path and config_path.exists():
            try:
                with open(config_path, 'r', encoding='utf-8') as f:
                    json_config = json.load(f)
            except Exception as e:
                print(f"Ошибка чтения конфига: {e}")
        else:
            pass

        cli_config = self._parse_cli_args()

        merged = {**default_config, **json_config, **cli_config}
        merged["rows_to_keep"] = self._validate_rows_param(merged["rows_to_keep"])

        return merged

    def _parse_cli_args(self) -> Dict[str, Any]:
        """Парсинг аргументов командной строки."""
        parser = argparse.ArgumentParser(description="Обрезка Excel/ODS файлов")
        parser.add_argument("--source", type=str, help="Папка с исходными файлами")
        parser.add_argument("--output", type=str, help="Папка для сохранения результатов")
        parser.add_argument("--rows", type=int, help="Количество строк для сохранения")
        parser.add_argument("--config", type=str, help="Путь к JSON-конфигу")
        parser.add_argument("--workers", type=int, help="Количество потоков")
        parser.add_argument("--preserve-formatting", action="store_true", help="Сохранять форматирование")

        args, _ = parser.parse_known_args()

        cli_config = {}
        if args.source:
            cli_config["source_folder"] = args.source
        if args.output:
            cli_config["output_folder"] = args.output
        if args.rows is not None:
            cli_config["rows_to_keep"] = args.rows
        if args.workers is not None:
            cli_config["max_workers"] = args.workers
        if args.preserve_formatting:
            cli_config["preserve_formatting"] = True

        return cli_config

    def _validate_rows_param(self, value: Any) -> int:
        """Валидация параметра rows_to_keep."""
        try:
            rows = int(value)
            if rows < 0:
                self.logger.warning(f"Отрицательное значение rows_to_keep ({rows}), установлено 0")
                return 0
            return rows
        except (ValueError, TypeError):
            self.logger.warning(f"Некорректное значение rows_to_keep ({value}), установлено 0")
            return 0

    def _resolve_path(self, path_str: str, relative_to: Optional[Path] = None) -> Path:
        """Преобразует строку пути в абсолютный Path."""
        if not path_str:
            return Path()
        path = Path(path_str).expanduser()
        if not path.is_absolute() and relative_to is not None:
            path = relative_to / path
        return path


    # def _setup_logging(self):
    #     """Настройка полноценного логирования."""
    #     log_format = "%(asctime)s [%(levelname)s] %(message)s"

    #     # log_file_path = Path(self.config["log_file"])
    #     if self.config["log_file"]:
    #         log_file_path_str = self.config["log_file"]
    #         log_file_path = Path(log_file_path_str)
    #         if not log_file_path.is_absolute():
    #             if log_file_path_str.startswith('~'):
    #                 # Заменяем '~' на путь к домашней директории
    #                 expand_path = os.path.expanduser(log_file_path_str)  # Это расширит '~' или '~/путь'
    #                 log_file_path = Path(expand_path)  # Создаем новый Path объект
    #             else:
    #                 log_file_path =  self.work_dir / Path(self.config["log_file"])   
    #         else:
    #             pass
    #     else:
    #         log_file_path =  self.work_dir / "cut_files.log"

    #     log_file_path.parent.mkdir(parents=True, exist_ok=True)
    #     logger = logging.getLogger()
    #     has_file = any(isinstance(h, logging.FileHandler) and h.baseFilename == str(log_file_path.resolve()) for h in logger.handlers)
    #     has_stream = any(isinstance(h, logging.FileHandler) for h in logger.handlers)

    #     if not has_file:
    #         fh = logging.FileHandler(log_file_path, encoding="utf-8")
    #         fh.setFormatter(logging.Formatter(log_format))
    #         logger.addHandler(fh)
    #     if not has_stream:
    #         sh = logging.StreamHandler(sys.stdout)
    #         sh.setFormatter(logging.Formatter(log_format))

    #     logger.setLevel(getattr(logging, self.config["log_level"].upper(), logging.INFO))
    def _setup_logging(self):
        log_format = "%(asctime)s [%(levelname)s] %(message)s"
        log_file_path = self._resolve_path(self.config["log_file"], self.work_dir)
        if not log_file_path.name:  # если путь пустой
            log_file_path = self.work_dir / "cut_files.log"
        log_file_path.parent.mkdir(parents=True, exist_ok=True)

        logger = logging.getLogger()
        logger.handlers.clear()  # очистка старых обработчиков

        fh = logging.FileHandler(log_file_path, encoding="utf-8")
        fh.setFormatter(logging.Formatter(log_format))
        logger.addHandler(fh)

        sh = logging.StreamHandler(sys.stdout)
        sh.setFormatter(logging.Formatter(log_format))
        logger.addHandler(sh)

        logger.setLevel(getattr(logging, self.config["log_level"].upper(), logging.INFO))

    # def read_all_ods_sheets(self, filepath: Path) -> Dict[str, pd.DataFrame]:
    #     """
    #     Чтение всех листов ODS-файла через engine='odf'.

    #     Args:
    #         filepath: Путь к файлу .ods.

    #     Returns:
    #         Словарь {имя_листа: DataFrame}.
    #     """
    #     # try:
    #     #     sheets = pd.read_excel(filepath, sheet_name=None, engine="odf", header=None, dtype=object)
    #     #     return sheets
    #     # except Exception as e:
    #     #     self.logger.error(f"Ошибка чтения ODS {filepath.name}: {e}")
    #     #     return {}
        
    #     all_sheets = {}
    #     # 1. Используем ezodf для открытия документа и получения имен листов
    #     try:
    #         doc = ezodf.opendoc(filepath)
    #         sheet_names = [sheet.name for sheet in doc.sheets if len(sheet.name)<32]


    #     except Exception as e:
    #         print(f"Ошибка при открытии файла {filepath} для получения списка листов: {e}")
    #         return {}
    #     finally:
    #         pass

            
    #     for sheet_name in sheet_names:
    #         try:
    #             df = pd.read_excel(filepath, sheet_name=sheet_name, engine='calamine', header=None, dtype=object)
    #             all_sheets[sheet_name] = df
    #             # print(f"Лист '{sheet_name}' успешно загружен. Размер: {df.shape}")
    #         except Exception as e:
    #             print(f"Ошибка при загрузке листа '{sheet_name}': {e}")
    #             # Продолжаем загрузку остальных листов
    #             continue

    #     return all_sheets
    def read_all_ods_sheets(self, filepath: Path) -> Dict[str, pd.DataFrame]:
        all_sheets = {}
        try:
            doc = ezodf.opendoc(filepath)
            sheet_names = [sheet.name for sheet in doc.sheets if len(sheet.name) < 32]
        except Exception as e:
            self.logger.error(f"Ошибка при открытии файла {filepath} для получения списка листов: {e}")
            return {}

        for sheet_name in sheet_names:
            try:
                df = pd.read_excel(filepath, sheet_name=sheet_name, engine='calamine', header=None, dtype=object)
                all_sheets[sheet_name] = df
            except Exception as e:
                self.logger.error(f"Ошибка при загрузке листа '{sheet_name}': {e}")
                continue
        return all_sheets

    def get_file_stats(self, sheets: Dict[str, pd.DataFrame]) -> Dict[str, int]:
        """Подсчёт исходного количества строк по листам."""
        return {name: len(df) for name, df in sheets.items()}

    def process_single_file(self, filepath: Path, output_folder: Path) -> Optional[Dict[str, Any]]:
        """
        Обработка одного файла.

        Returns:
            Словарь со статистикой или None при ошибке.
        """
        start_time = datetime.now()
        file_stats = {
            "file_name": filepath.name,
            "source_path": str(filepath),
            "status": "success",
            "error_message": ""
        }
        self.logger.debug(f"Начало обработки файла: {filepath}")
        try:
            if filepath.suffix.lower() == ".ods":
                sheets = self.read_all_ods_sheets(filepath)
            else:
                engine = "calamine" if filepath.suffix.lower() == ".xlsb" else "openpyxl"
                sheets = pd.read_excel(filepath, sheet_name=None, engine=engine, header=None, dtype=object)

            original_rows = self.get_file_stats(sheets)
            file_stats["original_rows"] = sum(original_rows.values())
            file_stats["sheets_count"] = len(sheets)
            self.logger.debug(f"Количество листов в файле: {len(sheets)}")
            rows_to_keep = self.config["rows_to_keep"]
            cut_sheets = {}
            for sheet_name, df in sheets.items():
                if rows_to_keep > 0:
                    cut_sheets[sheet_name] = df.iloc[:rows_to_keep].copy()
                else:
                    cut_sheets[sheet_name] = df.copy()

            relative_path = filepath.relative_to(self.source_dir)
            output_path = output_folder / relative_path.with_suffix(".xlsx")
            output_path.parent.mkdir(parents=True, exist_ok=True)

            if self.config["preserve_formatting"]:
                self._save_with_formatting(cut_sheets, filepath, output_path)
            else:
                with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                    for sheet_name, df_cut in cut_sheets.items():
                        safe_name = sheet_name[:31]
                        df_cut.to_excel(writer, sheet_name=safe_name, header=False, index=False)

            file_stats["output_path"] = str(output_path)
            file_stats["rows_saved"] = sum(len(df) for df in cut_sheets.values())
            file_stats["processing_time_sec"] = (datetime.now() - start_time).total_seconds()

            self.logger.info(f"Обработан {filepath.name}: {file_stats['original_rows']} -> {file_stats['rows_saved']} строк")

        except Exception as e:
            file_stats["status"] = "error"
            file_stats["error_message"] = str(e)
            self.logger.error(f"Ошибка обработки {filepath.name}: {e}")

        return file_stats

    def _save_with_formatting(self, cut_sheets: Dict[str, pd.DataFrame], source_path: Path, output_path: Path):
        """
        Сохранение с сохранением форматирования путём удаления лишних строк.
        """
        if source_path.suffix.lower() not in [".xlsx", ".xlsm"]:
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for sheet_name, df_cut in cut_sheets.items():
                    df_cut.to_excel(writer, sheet_name=sheet_name[:31], header=False, index=False)
            return

        try:
            from openpyxl import load_workbook

            wb = load_workbook(source_path)
            rows_to_keep = self.config["rows_to_keep"]

            for sheet_name in list(wb.sheetnames):
                if sheet_name not in cut_sheets:
                    del wb[sheet_name]
                elif rows_to_keep > 0:
                    ws = wb[sheet_name]
                    if ws.max_row > rows_to_keep:
                        ws.delete_rows(rows_to_keep + 1, ws.max_row - rows_to_keep)

            wb.save(output_path)

        except Exception as e:
            self.logger.warning(f"Не удалось сохранить форматирование: {e}. Сохранение без форматирования.")
            with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
                for sheet_name, df_cut in cut_sheets.items():
                    df_cut.to_excel(writer, sheet_name=sheet_name[:31], header=False, index=False)

    def save_stats_to_excel(self):
        """Сохранение статистики в Excel-файл."""
        if not self.stats:
            return

        stats_df = pd.DataFrame(self.stats)
        if self.config["stats_file"]:
            stats_path_str = self.config["stats_file"]
            stats_path = Path(stats_path_str)
            if not stats_path.is_absolute():
                if stats_path_str.startswith('~'):
                    # Заменяем '~' на путь к домашней директории
                    expanded_path = os.path.expanduser(stats_path_str)  # Это расширит '~' или '~/путь'
                    stats_path = Path(expanded_path)  # Создаем новый Path объект
                else:
                    stats_path = self.work_dir / self.config["stats_file"]
            else:
                pass
        else:
            stats_path = self.work_dir / "processing_stats.xlsx"

        stats_path.parent.mkdir(parents=True, exist_ok=True)

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        stats_path = stats_path.with_stem(f"{stats_path.stem}_{timestamp}")

        success_count = sum(1 for s in self.stats if s["status"] == "success")
        error_count = len(self.stats) - success_count
        total_original = sum(s.get("original_rows", 0) for s in self.stats)
        total_saved = sum(s.get("rows_saved", 0) for s in self.stats)
        total_time = sum(s.get("processing_time_sec", 0) for s in self.stats)

        summary_row = pd.DataFrame([{
            "file_name": "ИТОГО",
            "source_path": "",
            "status": f"Успешно: {success_count}, Ошибок: {error_count}",
            "error_message": "",
            "original_rows": total_original,
            "sheets_count": "",
            "output_path": "",
            "rows_saved": total_saved,
            "processing_time_sec": total_time
        }])

        stats_df = pd.concat([stats_df, summary_row], ignore_index=True)

        with pd.ExcelWriter(stats_path, engine="openpyxl") as writer:
            stats_df.to_excel(writer, sheet_name="Детализация", index=False)

        self.logger.info(f"Статистика сохранена в {stats_path}")

    def process_files(self):
        """Основной метод обработки."""
        self.logger.info("=" * 60)
        self.logger.info("Запуск скрипта обрезки Excel/ODS файлов")
        self.logger.info(f"Рабочая директория: {self.work_dir}")
        self.logger.info(f"Конфигурация: {self.config}")
        self.logger.info("=" * 60)

        # if self.config["source_folder"]:
        #     source_folder_str = self.config["source_folder"]
        #     source_folder = Path(source_folder_str)
        #     if not source_folder.is_absolute():
        #         if source_folder_str.startswith('~'):
        #             # Заменяем '~' на путь к домашней директории
        #             expand_path = os.path.expanduser(source_folder_str)  # Это расширит '~' или '~/путь'
        #             source_folder = Path(expand_path)  # Создаем новый Path объект
        #             self.source_dir = source_folder
        #         else:
        #             self.source_dir = self.work_dir / self.config["source_folder"]
        #     else:
        #         self.source_dir = source_folder
        # else:        
        #     self.source_dir = self.work_dir / self.config["source_folder"]

        # if not self.source_dir.is_dir():
        #     self.logger.error(f"Папка {self.source_dir} не найдена")
        #     return

        # if self.config["output_folder"]:
        #     output_folder_str = self.config["output_folder"]
        #     output_folder = Path(output_folder_str )
        #     if not output_folder.is_absolute():
        #         if output_folder_str.startswith('~'):
        #             # Заменяем '~' на путь к домашней директории
        #             expanded_path = os.path.expanduser(output_folder_str)  # Это расширит '~' или '~/путь'
        #             output_folder = Path(expanded_path)  # Создаем новый Path объект
        #         else:
        #             output_folder = self.work_dir / output_folder    
        # else:
        #     output_folder = self.source_dir.parent / (self.source_dir.name + "_cut")

        # output_folder.mkdir(parents=True, exist_ok=True)
        # self.logger.info(f"Выходная папка: {output_folder}")
        self.source_dir = self._resolve_path(self.config["source_folder"], self.work_dir)
        if not self.source_dir.is_dir():
            self.logger.error(f"Папка {self.source_dir} не найдена")
            return

        output_folder = self._resolve_path(self.config["output_folder"], self.work_dir) if self.config["output_folder"] else self.source_dir.parent / (self.source_dir.name + "_cut")
        output_folder.mkdir(parents=True, exist_ok=True)
        
        extensions = ["*.xlsb", "*.xlsx", "*.xlsm", "*.ods"]
        all_files = []
        for ext in extensions:
            all_files.extend(f for f in self.source_dir.rglob(ext) if not f.name.startswith("~"))

        if not all_files:
            self.logger.warning("Файлы для обработки не найдены")
            return

        ext_counts = Counter(f.suffix.lower() for f in all_files)
        self.logger.info(f"Найдено файлов: {len(all_files)} (xlsb: {ext_counts['.xlsb']}, xlsx: {ext_counts['.xlsx']}, xlsm: {ext_counts['.xlsm']}, ods: {ext_counts['.ods']})")

        total_start = datetime.now()
        max_workers = self.config["max_workers"]

        if max_workers > 1:
            with ThreadPoolExecutor(max_workers=max_workers) as executor:
                futures = {executor.submit(self.process_single_file, f, output_folder): f for f in all_files}
                for future in as_completed(futures):
                    result = future.result()
                    if result:
                        self.stats.append(result)
        else:
            for idx, filepath in enumerate(all_files, 1):
                self.logger.info(f"[{idx}/{len(all_files)}] Обработка {filepath.name}")
                result = self.process_single_file(filepath, output_folder)
                if result:
                    self.stats.append(result)

        total_time = (datetime.now() - total_start).total_seconds()

        success_count = sum(1 for s in self.stats if s["status"] == "success")
        error_count = len(self.stats) - success_count
        total_rows_original = sum(s.get("original_rows", 0) for s in self.stats)
        total_rows_saved = sum(s.get("rows_saved", 0) for s in self.stats)

        self.logger.info("=" * 60)
        self.logger.info("Обработка завершена")
        self.logger.info(f"Всего файлов: {len(all_files)}")
        self.logger.info(f"Успешно: {success_count}")
        self.logger.info(f"С ошибками: {error_count}")
        self.logger.info(f"Исходное количество строк: {total_rows_original}")
        self.logger.info(f"Сохранено строк: {total_rows_saved}")
        self.logger.info(f"Общее время: {total_time:.2f} сек")
        self.logger.info("=" * 60)

        self.save_stats_to_excel()
