import sys
import pandas as pd
from pathlib import Path
from PySide6.QtWidgets import (
    QApplication, QMainWindow, QWidget, QVBoxLayout, QHBoxLayout,
    QPushButton, QLabel, QFileDialog, QTableWidget, QTableWidgetItem,
    QComboBox, QListWidget, QListWidgetItem, QDialog, QMessageBox,
    QSpinBox, QGroupBox, QFormLayout, QCheckBox, QScrollArea,
    QPushButton as QPushButtonType
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QColor, QBrush, QFont
import re
from typing import Optional, Tuple, List, Dict, Any

# ==================== МОДУЛЬ: excel_core (автоопределение шапки и первого столбца) ====================

DEFAULT_HEADER_SCAN_ROWS = 50
DEFAULT_PREVIEW_PAGE_SIZE = 50
HEADER_LOOKAHEAD_ROWS = 6
HEADER_AUTO_ACCEPT_THRESHOLD = 0.58
HEADER_MIN_SCORE = 6.0
SERVICE_WORD_PATTERNS = (
    "содержание", "описание", "итого", "раздел", "лист", "комментар",
    "summary", "description", "total", "section"
)


def detect_header_row_and_start_col(frame: pd.DataFrame) -> Dict[str, Any]:
    """
    Автоматическое определение строки заголовка И первого значимого столбца.
    Возвращает: header_row, start_col, confidence
    """
    if frame.empty:
        return {
            "header_row": 0,
            "start_col": 0,
            "confidence": 0.0,
            "auto_accept": False,
            "explanation": "Нет строк для анализа.",
        }
    
    # Сначала определяем строку заголовка
    header_detection = _detect_header_row_only(frame)
    header_row = header_detection["header_row"]
    
    # Затем определяем первый значимый столбец на основе найденной строки заголовка
    start_col = _detect_first_data_column(frame, header_row)
    
    confidence = header_detection["confidence"]
    auto_accept = header_detection["auto_accept"]
    
    return {
        "header_row": header_row,
        "start_col": start_col,
        "confidence": confidence,
        "auto_accept": auto_accept,
        "explanation": header_detection.get("explanation", ""),
    }


def _detect_header_row_only(frame: pd.DataFrame) -> Dict[str, Any]:
    """Только определение строки заголовка"""
    if frame.empty:
        return {
            "header_row": 0,
            "confidence": 0.0,
            "auto_accept": False,
            "explanation": "Нет строк для анализа.",
        }

    candidates = []
    for row_index in range(min(len(frame.index), DEFAULT_HEADER_SCAN_ROWS)):
        features = _extract_row_features(frame, row_index)
        header_score, header_reasons = _score_header_likeness(features)
        data_score, metrics, data_reasons = _score_data_below(frame, row_index, features)
        penalty, anti_reasons = _score_anti_patterns(features)
        final_score = header_score + data_score - penalty
        
        candidates.append({
            "row_index": row_index,
            "final_score": round(final_score, 3),
            "header_reasons": header_reasons,
            "data_reasons": data_reasons,
            "anti_reasons": anti_reasons,
        })

    ordered = sorted(candidates, key=lambda x: (x["final_score"], -x["row_index"]), reverse=True)
    best = ordered[0]
    
    second_score = ordered[1]["final_score"] if len(ordered) > 1 else best["final_score"] - 5.0
    score_gap = float(best["final_score"] - second_score)
    strength = max(0.0, min(1.0, best["final_score"] / 18.0))
    separation = max(0.0, min(1.0, score_gap / 6.0))
    confidence = round((strength * 0.6) + (separation * 0.4), 3)
    auto_accept = (
        best["final_score"] >= HEADER_MIN_SCORE
        and confidence >= HEADER_AUTO_ACCEPT_THRESHOLD
    )
    
    return {
        "header_row": int(best["row_index"]),
        "confidence": confidence,
        "auto_accept": auto_accept,
        "explanation": ", ".join(best["header_reasons"] + best["data_reasons"]),
    }


def _detect_first_data_column(frame: pd.DataFrame, header_row: int) -> int:
    """
    Определение первого столбца, содержащего данные.
    Анализирует строку заголовка и несколько строк под ней.
    """
    if frame.empty or header_row >= len(frame.index):
        return 0
    
    # Получаем строку заголовка
    header_row_data = frame.iloc[header_row]
    
    # Определяем количество строк для анализа под заголовком
    max_check_rows = min(10, len(frame.index) - header_row - 1)
    if max_check_rows == 0:
        return 0
    
    start_col = 0
    
    # Анализируем каждый столбец, начиная с первого
    for col_idx in range(len(frame.columns)):
        # Проверяем, пустой ли заголовок в этом столбце
        header_value = header_row_data.iloc[col_idx]
        header_is_empty = pd.isna(header_value) or str(header_value).strip() == ''
        
        # Если заголовок пустой, вероятно, это не начало таблицы
        if header_is_empty:
            # Но проверим, может быть данные ниже не пустые?
            has_data_below = False
            for row_offset in range(1, max_check_rows + 1):
                row_idx = header_row + row_offset
                if row_idx < len(frame.index):
                    cell_value = frame.iloc[row_idx, col_idx]
                    if pd.notna(cell_value) and str(cell_value).strip() != '':
                        has_data_below = True
                        break
            
            if not has_data_below:
                start_col = col_idx + 1
                continue
            else:
                # Есть данные ниже, значит это начало таблицы
                return col_idx
        else:
            # Заголовок не пустой - это начало таблицы
            return col_idx
    
    return start_col


def _extract_row_features(frame: pd.DataFrame, row_index: int) -> Dict[str, Any]:
    """Извлечение признаков строки для анализа"""
    row = frame.iloc[row_index].tolist()
    labels = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
    
    text_like_count = sum(1 for v in labels if any(c.isalpha() for c in v))
    numeric_like_count = sum(1 for v in labels if _looks_numeric(v))
    short_label_count = sum(1 for v in labels if len(v) <= 32 and len(v.split()) <= 5)
    long_phrase_count = sum(1 for v in labels if len(v) >= 45 or len(v.split()) >= 6)
    service_word_hits = sum(_service_word_hits(v) for v in labels)
    
    return {
        "filled_count": len(labels),
        "text_like_count": text_like_count,
        "numeric_like_count": numeric_like_count,
        "short_label_count": short_label_count,
        "long_phrase_count": long_phrase_count,
        "service_word_hits": service_word_hits,
    }


def _looks_numeric(text: str) -> bool:
    return bool(re.fullmatch(r"-?\d+([.,]\d+)?", text.strip()))


def _service_word_hits(text: str) -> int:
    lowered = text.casefold()
    return sum(1 for pattern in SERVICE_WORD_PATTERNS if pattern in lowered)


def _score_header_likeness(features: Dict[str, Any]) -> Tuple[float, List[str]]:
    if features["filled_count"] == 0:
        return -10.0, ["пустая строка"]
    
    score = 0.0
    reasons = []
    
    score += features["filled_count"] * 1.4
    score += features["text_like_count"] * 1.2
    score += features["short_label_count"] * 0.8
    
    if features["numeric_like_count"] >= max(2, features["filled_count"] - 1):
        score -= features["numeric_like_count"] * 1.8
    
    if features["long_phrase_count"]:
        score -= features["long_phrase_count"] * 2.5
    
    if features["service_word_hits"]:
        score -= features["service_word_hits"] * 2.8
    
    return score, reasons


def _score_data_below(
    frame: pd.DataFrame, row_index: int, features: Dict[str, Any]
) -> Tuple[float, Dict[str, Any], List[str]]:
    lookahead = min(len(frame.index), row_index + 1 + HEADER_LOOKAHEAD_ROWS)
    stable_rows = 0
    typed_value_hits = 0
    
    for next_row_index in range(row_index + 1, lookahead):
        row = frame.iloc[next_row_index].tolist()
        values = [str(v).strip() for v in row if pd.notna(v) and str(v).strip()]
        
        if not values:
            continue
        
        numeric_hits = sum(1 for v in values if _looks_numeric(v))
        if numeric_hits:
            typed_value_hits += numeric_hits
        
        if len(values) >= 3:
            stable_rows += 1
    
    score = stable_rows * 2.2 + min(4.0, typed_value_hits * 0.4)
    metrics = {"stable_rows": stable_rows, "typed_value_hits": typed_value_hits}
    reasons = ["ниже найден табличный блок"] if stable_rows >= 2 else []
    
    return score, metrics, reasons


def _score_anti_patterns(features: Dict[str, Any]) -> Tuple[float, List[str]]:
    penalty = 0.0
    reasons = []
    
    if features["service_word_hits"]:
        penalty += features["service_word_hits"] * 1.6
    if features["long_phrase_count"]:
        penalty += features["long_phrase_count"] * 1.4
    
    return penalty, reasons


# ==================== КЛАСС: Thread для фоновой загрузки ====================

class LoadFileThread(QThread):
    finished = Signal(object, object, object, str, str)
    
    def __init__(self, file_path: str, sheet_name: str = None):
        super().__init__()
        self.file_path = file_path
        self.sheet_name = sheet_name
    
    def run(self):
        try:
            excel_file = pd.ExcelFile(self.file_path)
            sheets = excel_file.sheet_names
            
            if self.sheet_name is None:
                self.sheet_name = sheets[0] if sheets else None
            
            if self.sheet_name:
                df = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None)
                df_raw = pd.read_excel(self.file_path, sheet_name=self.sheet_name, header=None, dtype=str)
                self.finished.emit(df, df_raw, sheets, None, self.file_path)
            else:
                self.finished.emit(None, None, sheets, "Нет листов в файле", self.file_path)
        except Exception as e:
            self.finished.emit(None, None, None, str(e), self.file_path)


# ==================== КЛАСС: Диалог автоопределения шапки ====================

class AutoHeaderDialog(QDialog):
    def __init__(self, df: pd.DataFrame, parent=None):
        super().__init__(parent)
        self.df = df
        self.selected_row = 0
        self.selected_col = 0
        self.setWindowTitle("Автоопределение заголовка и начала данных")
        self.setModal(True)
        self.resize(900, 550)
        self._setup_ui()
        self._detect_and_display()
    
    def _setup_ui(self):
        layout = QVBoxLayout(self)
        
        # Информационная метка
        self.info_label = QLabel("Анализ структуры файла...")
        self.info_label.setWordWrap(True)
        layout.addWidget(self.info_label)
        
        # Таблица предпросмотра
        self.preview_table = QTableWidget()
        layout.addWidget(self.preview_table)
        
        # Кнопки
        btn_layout = QHBoxLayout()
        self.accept_btn = QPushButton("Принять")
        self.accept_btn.clicked.connect(self.accept)
        self.accept_btn.setEnabled(False)
        self.manual_btn = QPushButton("Выбрать вручную")
        self.manual_btn.clicked.connect(self.reject)
        
        btn_layout.addStretch()
        btn_layout.addWidget(self.manual_btn)
        btn_layout.addWidget(self.accept_btn)
        layout.addLayout(btn_layout)
    
    def _detect_and_display(self):
        """Запуск автоопределения и отображение результатов"""
        detection = detect_header_row_and_start_col(self.df)
        self.selected_row = detection["header_row"]
        self.selected_col = detection["start_col"]
        
        # Отображение таблицы
        self._display_preview()
        
        # Подсветка найденной строки и столбца
        if self.selected_row < self.preview_table.rowCount():
            for col in range(self.preview_table.columnCount()):
                item = self.preview_table.item(self.selected_row, col)
                if item:
                    item.setBackground(QBrush(QColor(255, 255, 150)))
        
        if self.selected_col < self.preview_table.columnCount():
            for row in range(self.preview_table.rowCount()):
                item = self.preview_table.item(row, self.selected_col)
                if item:
                    current_color = item.background().color()
                    if current_color == QColor(255, 255, 150):
                        item.setBackground(QBrush(QColor(255, 200, 100)))
                    else:
                        item.setBackground(QBrush(QColor(230, 240, 255)))
        
        confidence = detection["confidence"]
        auto_accept = detection["auto_accept"]
        explanation = detection["explanation"]
        
        self.info_label.setText(
            f"📊 Результаты автоопределения:\n"
            f"• Строка заголовка: **{self.selected_row + 1}**\n"
            f"• Первый столбец данных: **{self.selected_col + 1}**\n"
            f"• Уверенность: {confidence:.1%}\n"
            f"• Объяснение: {explanation}\n\n"
            f"{'✅ Рекомендуется принять' if auto_accept else '⚠️ Рекомендуется проверить вручную'}"
        )
        
        self.accept_btn.setEnabled(True)
    
    def _display_preview(self):
        """Отображение предпросмотра"""
        max_rows = min(30, len(self.df))
        max_cols = min(20, len(self.df.columns))
        
        self.preview_table.setRowCount(max_rows)
        self.preview_table.setColumnCount(max_cols)
        
        # Заголовки столбцов (номера)
        self.preview_table.setHorizontalHeaderLabels([f"{i+1}" for i in range(max_cols)])
        
        # Заголовки строк (номера)
        self.preview_table.setVerticalHeaderLabels([f"{i+1}" for i in range(max_rows)])
        
        for i in range(max_rows):
            for j in range(max_cols):
                value = self.df.iloc[i, j]
                item = QTableWidgetItem(str(value) if pd.notna(value) else "")
                self.preview_table.setItem(i, j, item)
        
        self.preview_table.resizeColumnsToContents()


# ==================== ОСНОВНОЙ КЛАСС ====================

class ExcelComparator(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Comparator - Сравнение файлов (с автоопределением)")
        self.setGeometry(100, 100, 1800, 900)
        
        # Переменные для хранения данных
        self.file1_path = None
        self.file2_path = None
        self.df1 = None
        self.df2 = None
        self.df1_raw = None
        self.df2_raw = None
        self.sheet1 = None
        self.sheet2 = None
        self.start_row1 = 0
        self.start_row2 = 0
        self.start_col1 = 0
        self.start_col2 = 0
        
        # Флаги для автоопределения
        self.header_auto_detected1 = False
        self.header_auto_detected2 = False
        
        # Потоки для фоновой загрузки
        self.load_thread1 = None
        self.load_thread2 = None
        
        self.init_ui()
    
    def init_ui(self):
        central_widget = QWidget()
        self.setCentralWidget(central_widget)
        main_layout = QVBoxLayout()
        
        # ==================== БЛОК 1: Загрузка файлов ====================
        files_group = QGroupBox("1. Загрузка файлов")
        files_layout = QHBoxLayout()
        
        self.label_file1 = QLabel("Файл 1: не загружен")
        btn_load_file1 = QPushButton("Загрузить файл 1")
        btn_load_file1.clicked.connect(lambda: self.load_file(1))
        
        self.label_file2 = QLabel("Файл 2: не загружен")
        btn_load_file2 = QPushButton("Загрузить файл 2")
        btn_load_file2.clicked.connect(lambda: self.load_file(2))
        
        files_layout.addWidget(btn_load_file1)
        files_layout.addWidget(self.label_file1)
        files_layout.addWidget(btn_load_file2)
        files_layout.addWidget(self.label_file2)
        files_group.setLayout(files_layout)
        main_layout.addWidget(files_group)
        
        # ==================== БЛОК 2: Выбор листов ====================
        sheets_group = QGroupBox("2. Выбор листов")
        sheets_layout = QHBoxLayout()
        
        sheets_layout.addWidget(QLabel("Лист файла 1:"))
        self.combo_sheet1 = QComboBox()
        self.combo_sheet1.currentTextChanged.connect(self.on_sheet1_changed)
        sheets_layout.addWidget(self.combo_sheet1)
        
        sheets_layout.addWidget(QLabel("Лист файла 2:"))
        self.combo_sheet2 = QComboBox()
        self.combo_sheet2.currentTextChanged.connect(self.on_sheet2_changed)
        sheets_layout.addWidget(self.combo_sheet2)
        
        sheets_group.setLayout(sheets_layout)
        main_layout.addWidget(sheets_group)
        
        # ==================== БЛОК 3: Выбор начала таблицы ====================
        start_group = QGroupBox("3. Выбор начала таблицы")
        start_layout = QVBoxLayout()
        
        # Кнопка автоопределения для файла 1
        header_row_layout = QHBoxLayout()
        header_row_layout.addWidget(QLabel("Файл 1:"))
        self.btn_auto_header1 = QPushButton("🔍 Автоопределение заголовка")
        self.btn_auto_header1.clicked.connect(lambda: self.auto_detect_header(1))
        self.btn_auto_header1.setEnabled(False)
        header_row_layout.addWidget(self.btn_auto_header1)
        
        header_row_layout.addWidget(QLabel("Файл 2:"))
        self.btn_auto_header2 = QPushButton("🔍 Автоопределение заголовка")
        self.btn_auto_header2.clicked.connect(lambda: self.auto_detect_header(2))
        self.btn_auto_header2.setEnabled(False)
        header_row_layout.addWidget(self.btn_auto_header2)
        header_row_layout.addStretch()
        start_layout.addLayout(header_row_layout)
        
        # Информация о выбранной позиции
        pos_layout = QHBoxLayout()
        pos_layout.addWidget(QLabel("Позиция в файле 1:"))
        self.label_start1 = QLabel("не выбрано")
        self.label_start1.setStyleSheet("background-color: #fff3cd; padding: 5px;")
        pos_layout.addWidget(self.label_start1)
        
        pos_layout.addWidget(QLabel("Позиция в файле 2:"))
        self.label_start2 = QLabel("не выбрано")
        self.label_start2.setStyleSheet("background-color: #fff3cd; padding: 5px;")
        pos_layout.addWidget(self.label_start2)
        pos_layout.addStretch()
        start_layout.addLayout(pos_layout)
        
        start_group.setLayout(start_layout)
        main_layout.addWidget(start_group)
        
        # ==================== БЛОК 4: Таблицы для выбора начала ====================
        tables_layout = QHBoxLayout()
        
        # Таблица файла 1
        left_layout = QVBoxLayout()
        left_layout.addWidget(QLabel("Предпросмотр файла 1 (двойной клик для выбора позиции):"))
        self.table_preview1 = QTableWidget()
        self.table_preview1.itemDoubleClicked.connect(lambda item: self.set_start_position(1, item))
        left_layout.addWidget(self.table_preview1)
        
        # Таблица файла 2
        right_layout = QVBoxLayout()
        right_layout.addWidget(QLabel("Предпросмотр файла 2 (двойной клик для выбора позиции):"))
        self.table_preview2 = QTableWidget()
        self.table_preview2.itemDoubleClicked.connect(lambda item: self.set_start_position(2, item))
        right_layout.addWidget(self.table_preview2)
        
        tables_layout.addLayout(left_layout)
        tables_layout.addLayout(right_layout)
        main_layout.addLayout(tables_layout)
        
        # ==================== БЛОК 5: Выбор параметров сравнения ====================
        params_group = QGroupBox("4. Выбор параметров сравнения")
        params_layout = QVBoxLayout()
        
        # Подблок: Ключевые столбцы
        key_layout = QHBoxLayout()
        key_layout.addWidget(QLabel("Ключевые столбцы (для поиска строк):"))
        key_layout.addWidget(QLabel("Файл 1:"))
        self.list_key_cols1 = QListWidget()
        self.list_key_cols1.setMaximumHeight(80)
        key_layout.addWidget(self.list_key_cols1)
        
        key_layout.addWidget(QLabel("Файл 2:"))
        self.list_key_cols2 = QListWidget()
        self.list_key_cols2.setMaximumHeight(80)
        key_layout.addWidget(self.list_key_cols2)
        
        params_layout.addLayout(key_layout)
        
        # Подблок: Столбцы для анализа
        analysis_layout = QHBoxLayout()
        analysis_layout.addWidget(QLabel("Столбцы для анализа расхождений:"))
        analysis_layout.addWidget(QLabel("Файл 1:"))
        self.list_analysis_cols1 = QListWidget()
        self.list_analysis_cols1.setMaximumHeight(80)
        analysis_layout.addWidget(self.list_analysis_cols1)
        
        analysis_layout.addWidget(QLabel("Файл 2:"))
        self.list_analysis_cols2 = QListWidget()
        self.list_analysis_cols2.setMaximumHeight(80)
        analysis_layout.addWidget(self.list_analysis_cols2)
        
        params_layout.addLayout(analysis_layout)
        params_group.setLayout(params_layout)
        main_layout.addWidget(params_group)
        
        # ==================== БЛОК 6: Кнопки действий ====================
        actions_group = QGroupBox("5. Запуск анализа")
        actions_layout = QHBoxLayout()
        
        btn_compare = QPushButton("Сравнить файлы")
        btn_compare.setStyleSheet("background-color: #28a745; color: white; font-weight: bold; padding: 10px;")
        btn_compare.clicked.connect(self.compare_files)
        
        btn_save = QPushButton("Сохранить результат")
        btn_save.setStyleSheet("background-color: #007bff; color: white; padding: 10px;")
        btn_save.clicked.connect(self.save_result)
        btn_save.setEnabled(False)
        self.btn_save = btn_save
        
        actions_layout.addWidget(btn_compare)
        actions_layout.addWidget(btn_save)
        actions_group.setLayout(actions_layout)
        main_layout.addWidget(actions_group)
        
        # ==================== БЛОК 7: Результаты ====================
        result_group = QGroupBox("Результаты сравнения")
        result_layout = QVBoxLayout()
        
        self.table_result = QTableWidget()
        result_layout.addWidget(self.table_result)
        
        result_group.setLayout(result_layout)
        main_layout.addWidget(result_group)
        
        central_widget.setLayout(main_layout)
    
    def load_file(self, file_num):
        """Загрузка Excel файла с фоновым потоком"""
        file_path, _ = QFileDialog.getOpenFileName(
            self, f"Загрузить файл {file_num}",
            "", "Excel файлы (*.xlsx *.xls *.xlsm);;Все файлы (*)"
        )
        
        if not file_path:
            return
        
        thread = LoadFileThread(file_path)
        
        if file_num == 1:
            self.load_thread1 = thread
            self.label_file1.setText(f"Файл 1: {Path(file_path).name} (загрузка...)")
            thread.finished.connect(self._on_file1_loaded)
        else:
            self.load_thread2 = thread
            self.label_file2.setText(f"Файл 2: {Path(file_path).name} (загрузка...)")
            thread.finished.connect(self._on_file2_loaded)
        
        thread.start()
    
    def _on_file1_loaded(self, df, df_raw, sheets, error, file_path):
        if error:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке файла 1: {error}")
            self.label_file1.setText("Файл 1: ошибка загрузки")
            return
        
        self.file1_path = file_path
        self.df1 = df
        self.df1_raw = df_raw
        
        self.label_file1.setText(f"Файл 1: {Path(file_path).name}")
        
        self.combo_sheet1.blockSignals(True)
        self.combo_sheet1.clear()
        self.combo_sheet1.addItems(sheets)
        self.combo_sheet1.blockSignals(False)
        
        if sheets:
            self.sheet1 = sheets[0]
            self.update_preview1()
            self.btn_auto_header1.setEnabled(True)
        
        QMessageBox.information(self, "Успех", f"Файл 1 загружен!\nЛистов: {len(sheets)}")
    
    def _on_file2_loaded(self, df, df_raw, sheets, error, file_path):
        if error:
            QMessageBox.critical(self, "Ошибка", f"Ошибка при загрузке файла 2: {error}")
            self.label_file2.setText("Файл 2: ошибка загрузки")
            return
        
        self.file2_path = file_path
        self.df2 = df
        self.df2_raw = df_raw
        
        self.label_file2.setText(f"Файл 2: {Path(file_path).name}")
        
        self.combo_sheet2.blockSignals(True)
        self.combo_sheet2.clear()
        self.combo_sheet2.addItems(sheets)
        self.combo_sheet2.blockSignals(False)
        
        if sheets:
            self.sheet2 = sheets[0]
            self.update_preview2()
            self.btn_auto_header2.setEnabled(True)
        
        QMessageBox.information(self, "Успех", f"Файл 2 загружен!\nЛистов: {len(sheets)}")
    
    def on_sheet1_changed(self):
        if self.file1_path and self.combo_sheet1.currentText():
            self.sheet1 = self.combo_sheet1.currentText()
            self.df1 = pd.read_excel(self.file1_path, sheet_name=self.sheet1, header=None)
            self.df1_raw = pd.read_excel(self.file1_path, sheet_name=self.sheet1, header=None, dtype=str)
            self.update_preview1()
            self.update_columns_list1()
            self.header_auto_detected1 = False
    
    def on_sheet2_changed(self):
        if self.file2_path and self.combo_sheet2.currentText():
            self.sheet2 = self.combo_sheet2.currentText()
            self.df2 = pd.read_excel(self.file2_path, sheet_name=self.sheet2, header=None)
            self.df2_raw = pd.read_excel(self.file2_path, sheet_name=self.sheet2, header=None, dtype=str)
            self.update_preview2()
            self.update_columns_list2()
            self.header_auto_detected2 = False
    
    def auto_detect_header(self, file_num: int):
        """Автоопределение строки заголовка и первого столбца"""
        df = self.df1 if file_num == 1 else self.df2
        
        if df is None:
            QMessageBox.warning(self, "Ошибка", "Сначала загрузите файл!")
            return
        
        dialog = AutoHeaderDialog(df, self)
        if dialog.exec() == QDialog.Accepted:
            header_row = dialog.selected_row
            start_col = dialog.selected_col
            
            if file_num == 1:
                self.start_row1 = header_row
                self.start_col1 = start_col
                self.label_start1.setText(f"Строка {header_row + 1}, Столбец {start_col + 1} (автоопределено)")
                self.header_auto_detected1 = True
                self.update_columns_list1()
                self._highlight_position_in_preview(self.table_preview1, header_row, start_col)
            else:
                self.start_row2 = header_row
                self.start_col2 = start_col
                self.label_start2.setText(f"Строка {header_row + 1}, Столбец {start_col + 1} (автоопределено)")
                self.header_auto_detected2 = True
                self.update_columns_list2()
                self._highlight_position_in_preview(self.table_preview2, header_row, start_col)
            
            QMessageBox.information(
                self, 
                "Успех", 
                f"Начало таблицы определено:\n"
                f"• Строка заголовка: {header_row + 1}\n"
                f"• Первый столбец: {start_col + 1}\n\n"
                f"Теперь выберите ключевые столбцы и столбцы для анализа."
            )
    
    def _highlight_position_in_preview(self, table: QTableWidget, row: int, col: int):
        """Подсветка строки заголовка и первого столбца в таблице предпросмотра"""
        # Подсветка всей строки заголовка
        for c in range(table.columnCount()):
            item = table.item(row, c)
            if item:
                item.setBackground(QBrush(QColor(255, 255, 150)))
        
        # Подсветка первого столбца данных
        for r in range(table.rowCount()):
            item = table.item(r, col)
            if item:
                current_color = item.background().color()
                if current_color == QColor(255, 255, 150):
                    item.setBackground(QBrush(QColor(255, 200, 100)))
                else:
                    item.setBackground(QBrush(QColor(230, 240, 255)))
    
    def update_preview1(self):
        if self.df1 is not None:
            max_rows = min(50, len(self.df1))
            max_cols = min(20, len(self.df1.columns))
            
            self.table_preview1.setRowCount(max_rows)
            self.table_preview1.setColumnCount(max_cols)
            
            self.table_preview1.setHorizontalHeaderLabels([f"{i+1}" for i in range(max_cols)])
            self.table_preview1.setVerticalHeaderLabels([f"{i+1}" for i in range(max_rows)])
            
            for i in range(max_rows):
                for j in range(max_cols):
                    value = self.df1.iloc[i, j]
                    item = QTableWidgetItem(str(value) if pd.notna(value) else "")
                    self.table_preview1.setItem(i, j, item)
            
            self.table_preview1.resizeColumnsToContents()
            
            if self.header_auto_detected1 and self.start_row1 < max_rows:
                self._highlight_position_in_preview(self.table_preview1, self.start_row1, self.start_col1)
    
    def update_preview2(self):
        if self.df2 is not None:
            max_rows = min(50, len(self.df2))
            max_cols = min(20, len(self.df2.columns))
            
            self.table_preview2.setRowCount(max_rows)
            self.table_preview2.setColumnCount(max_cols)
            
            self.table_preview2.setHorizontalHeaderLabels([f"{i+1}" for i in range(max_cols)])
            self.table_preview2.setVerticalHeaderLabels([f"{i+1}" for i in range(max_rows)])
            
            for i in range(max_rows):
                for j in range(max_cols):
                    value = self.df2.iloc[i, j]
                    item = QTableWidgetItem(str(value) if pd.notna(value) else "")
                    self.table_preview2.setItem(i, j, item)
            
            self.table_preview2.resizeColumnsToContents()
            
            if self.header_auto_detected2 and self.start_row2 < max_rows:
                self._highlight_position_in_preview(self.table_preview2, self.start_row2, self.start_col2)
    
    def set_start_position(self, file_num, item):
        """Установка начальной позиции таблицы (ручной выбор)"""
        row = item.row()
        col = item.column()
        
        if file_num == 1:
            self.start_row1 = row
            self.start_col1 = col
            self.label_start1.setText(f"Строка {row + 1}, Столбец {col + 1} (вручную)")
            self.header_auto_detected1 = False
            self.update_columns_list1()
        else:
            self.start_row2 = row
            self.start_col2 = col
            self.label_start2.setText(f"Строка {row + 1}, Столбец {col + 1} (вручную)")
            self.header_auto_detected2 = False
            self.update_columns_list2()
    
    def update_columns_list1(self):
        """Обновление списка столбцов файла 1 (начиная с выбранного столбца)"""
        self.list_key_cols1.clear()
        self.list_analysis_cols1.clear()
        
        if self.df1 is not None and self.start_row1 < len(self.df1):
            header_row = self.df1.iloc[self.start_row1]
            # Начинаем с self.start_col1, а не с 0!
            for col_idx in range(self.start_col1, len(self.df1.columns)):
                col_name = str(header_row.iloc[col_idx]) if col_idx < len(header_row) else f"Столбец {col_idx + 1}"
                
                # Ключевые столбцы
                item_key = QListWidgetItem(f"{col_idx + 1}. {col_name}")
                item_key.setData(Qt.UserRole, col_idx)
                item_key.setCheckState(Qt.Unchecked)
                self.list_key_cols1.addItem(item_key)
                
                # Столбцы анализа
                item_analysis = QListWidgetItem(f"{col_idx + 1}. {col_name}")
                item_analysis.setData(Qt.UserRole, col_idx)
                item_analysis.setCheckState(Qt.Unchecked)
                self.list_analysis_cols1.addItem(item_analysis)
    
    def update_columns_list2(self):
        """Обновление списка столбцов файла 2 (начиная с выбранного столбца)"""
        self.list_key_cols2.clear()
        self.list_analysis_cols2.clear()
        
        if self.df2 is not None and self.start_row2 < len(self.df2):
            header_row = self.df2.iloc[self.start_row2]
            # Начинаем с self.start_col2, а не с 0!
            for col_idx in range(self.start_col2, len(self.df2.columns)):
                col_name = str(header_row.iloc[col_idx]) if col_idx < len(header_row) else f"Столбец {col_idx + 1}"
                
                # Ключевые столбцы
                item_key = QListWidgetItem(f"{col_idx + 1}. {col_name}")
                item_key.setData(Qt.UserRole, col_idx)
                item_key.setCheckState(Qt.Unchecked)
                self.list_key_cols2.addItem(item_key)
                
                # Столбцы анализа
                item_analysis = QListWidgetItem(f"{col_idx + 1}. {col_name}")
                item_analysis.setData(Qt.UserRole, col_idx)
                item_analysis.setCheckState(Qt.Unchecked)
                self.list_analysis_cols2.addItem(item_analysis)
    
    def format_number_with_comma(self, value):
        """Форматирование числа с запятой вместо точки"""
        if value is None:
            return "-"
        
        try:
            # Если это число, форматируем с запятой
            if isinstance(value, (int, float)):
                # Для целых чисел - без десятичной части
                if value == int(value):
                    return f"{int(value):,}".replace(",", " ")
                else:
                    # Для дробных - с запятой
                    return f"{value:.10f}".rstrip('0').rstrip('.').replace('.', ',')
            
            # Если это строка, пробуем конвертировать
            str_val = str(value).strip()
            # Заменяем точку на запятую для отображения
            return str_val.replace('.', ',')
        except:
            return str(value) if value is not None else "-"
    
    def try_convert_to_number(self, value):
        """Попытка конвертировать значение в число для сравнения (поддерживает запятую)"""
        if pd.isna(value) or value == '' or value == 'nan':
            return None
        
        try:
            if isinstance(value, (int, float)):
                return float(value)
            
            str_val = str(value).strip()
            # Заменяем запятую на точку для вычисления
            str_val_calc = str_val.replace(',', '.')
            return float(str_val_calc)
        except (ValueError, TypeError):
            return None
    
    def compare_values(self, val1_raw, val2_raw, val1_num, val2_num):
        """Сравнение значений с попыткой числового сравнения"""
        # Форматируем для отображения с запятой
        display_val1 = self.format_number_with_comma(val1_raw) if pd.notna(val1_raw) else ''
        display_val2 = self.format_number_with_comma(val2_raw) if pd.notna(val2_raw) else ''
        
        if val1_num is not None and val2_num is not None:
            return display_val1, display_val2, val1_num == val2_num, val1_num, val2_num
        else:
            # Строковое сравнение (очищаем от лишних пробелов)
            str1 = str(val1_raw).strip() if pd.notna(val1_raw) else ''
            str2 = str(val2_raw).strip() if pd.notna(val2_raw) else ''
            return display_val1, display_val2, str1 == str2, None, None
    
    def compare_files(self):
        """Сравнение файлов с использованием ключевых столбцов"""
        
        # Проверка ключевых столбцов
        key_cols1 = [self.list_key_cols1.item(i).data(Qt.UserRole) 
                     for i in range(self.list_key_cols1.count()) 
                     if self.list_key_cols1.item(i).checkState() == Qt.Checked]
        
        key_cols2 = [self.list_key_cols2.item(i).data(Qt.UserRole) 
                     for i in range(self.list_key_cols2.count()) 
                     if self.list_key_cols2.item(i).checkState() == Qt.Checked]
        
        if not key_cols1 or not key_cols2:
            QMessageBox.warning(self, "Ошибка", "Выберите ключевые столбцы для поиска!")
            return
        
        if len(key_cols1) != len(key_cols2):
            QMessageBox.warning(self, "Ошибка", "Количество ключевых столбцов должно быть одинаковым!")
            return
        
        # Проверка столбцов анализа
        analysis_cols1 = [self.list_analysis_cols1.item(i).data(Qt.UserRole) 
                         for i in range(self.list_analysis_cols1.count()) 
                         if self.list_analysis_cols1.item(i).checkState() == Qt.Checked]
        
        analysis_cols2 = [self.list_analysis_cols2.item(i).data(Qt.UserRole) 
                         for i in range(self.list_analysis_cols2.count()) 
                         if self.list_analysis_cols2.item(i).checkState() == Qt.Checked]
        
        if not analysis_cols1 or not analysis_cols2:
            QMessageBox.warning(self, "Ошибка", "Выберите столбцы для анализа расхождений!")
            return
        
        if len(analysis_cols1) != len(analysis_cols2):
            QMessageBox.warning(self, "Ошибка", "Количество столбцов анализа должно быть одинаковым!")
            return
        
        try:
            # Получение данных (начиная с выбранной позиции)
            data1_raw = self.df1_raw.iloc[self.start_row1 + 1:].copy()
            data2_raw = self.df2_raw.iloc[self.start_row2 + 1:].copy()
            
            data1_num = self.df1.iloc[self.start_row1 + 1:].copy()
            data2_num = self.df2.iloc[self.start_row2 + 1:].copy()
            
            # Удаление полностью пустых строк
            data1_raw = data1_raw.dropna(how='all').reset_index(drop=True)
            data2_raw = data2_raw.dropna(how='all').reset_index(drop=True)
            data1_num = data1_num.dropna(how='all').reset_index(drop=True)
            data2_num = data2_num.dropna(how='all').reset_index(drop=True)
            
            # Заголовки (используем выбранные столбцы)
            key_headers1 = [str(self.df1.iloc[self.start_row1, col]) for col in key_cols1]
            key_headers2 = [str(self.df2.iloc[self.start_row2, col]) for col in key_cols2]
            analysis_headers1 = [str(self.df1.iloc[self.start_row1, col]) for col in analysis_cols1]
            analysis_headers2 = [str(self.df2.iloc[self.start_row2, col]) for col in analysis_cols2]
            
            # Создание словарей для быстрого поиска
            dict2 = {}
            for idx, row in data2_raw.iterrows():
                key_values = tuple(str(row.iloc[col]).strip() if pd.notna(row.iloc[col]) else '' 
                                 for col in key_cols2)
                dict2[key_values] = {
                    'index': idx,
                    'analysis_values_raw': [data2_raw.iloc[idx, col] for col in analysis_cols2],
                    'analysis_values_num': [data2_num.iloc[idx, col] for col in analysis_cols2]
                }
            
            # Сравнение
            comparison_results = []
            found_in_file2 = set()
            
            for idx1, row1 in data1_raw.iterrows():
                key_values = tuple(str(row1.iloc[col]).strip() if pd.notna(row1.iloc[col]) else '' 
                                 for col in key_cols1)
                
                analysis_values1_raw = [data1_raw.iloc[idx1, col] for col in analysis_cols1]
                analysis_values1_num = [data1_num.iloc[idx1, col] for col in analysis_cols1]
                
                if key_values in dict2:
                    found_in_file2.add(key_values)
                    analysis_values2_raw = dict2[key_values]['analysis_values_raw']
                    analysis_values2_num = dict2[key_values]['analysis_values_num']
                    
                    for i, (val1_raw, val2_raw, val1_num, val2_num) in enumerate(
                        zip(analysis_values1_raw, analysis_values2_raw, analysis_values1_num, analysis_values2_num)):
                        
                        num1 = self.try_convert_to_number(val1_num)
                        num2 = self.try_convert_to_number(val2_num)
                        
                        display_val1, display_val2, is_equal, num1, num2 = self.compare_values(
                            val1_raw, val2_raw, num1, num2)
                        
                        if num1 is not None and num2 is not None:
                            difference = num2 - num1
                            # Форматируем разницу с запятой
                            if difference == int(difference):
                                difference_str = str(int(difference))
                            else:
                                difference_str = f"{difference:.10f}".rstrip('0').rstrip('.').replace('.', ',')
                        else:
                            difference_str = "-"
                        
                        row_data = {
                            "Ключ": " | ".join(key_values),
                            "Параметр": analysis_headers1[i],
                            "Значение в файле 1": display_val1,
                            "Значение в файле 2": display_val2,
                            "Разница": difference_str,
                            "Статус": "ОК" if is_equal else "РАСХОЖДЕНИЕ"
                        }
                        comparison_results.append(row_data)
                else:
                    for i, val1_raw in enumerate(analysis_values1_raw):
                        display_val1 = self.format_number_with_comma(val1_raw) if pd.notna(val1_raw) else ''
                        
                        row_data = {
                            "Ключ": " | ".join(key_values),
                            "Параметр": analysis_headers1[i],
                            "Значение в файле 1": display_val1,
                            "Значение в файле 2": "ОТСУТСТВУЕТ",
                            "Разница": "-",
                            "Статус": "ОТСУТСТВУЕТ В ФАЙЛЕ 2"
                        }
                        comparison_results.append(row_data)
            
            for key_values, data_info in dict2.items():
                if key_values not in found_in_file2:
                    analysis_values2_raw = data_info['analysis_values_raw']
                    for i, val2_raw in enumerate(analysis_values2_raw):
                        display_val2 = self.format_number_with_comma(val2_raw) if pd.notna(val2_raw) else ''
                        
                        row_data = {
                            "Ключ": " | ".join(key_values),
                            "Параметр": analysis_headers2[i],
                            "Значение в файле 1": "ОТСУТСТВУЕТ",
                            "Значение в файле 2": display_val2,
                            "Разница": "-",
                            "Статус": "ОТСУТСТВУЕТ В ФАЙЛЕ 1"
                        }
                        comparison_results.append(row_data)
            
            self.result_df = pd.DataFrame(comparison_results)
            self.display_results()
            self.btn_save.setEnabled(True)
            
            QMessageBox.information(self, "Успех", f"Сравнение завершено!\nНайдено {len(comparison_results)} записей.")
            
        except Exception as e:
            import traceback
            QMessageBox.critical(self, "Ошибка при сравнении", f"{str(e)}\n{traceback.format_exc()}")
    
    def display_results(self):
        """Отображение результатов"""
        self.table_result.setRowCount(len(self.result_df))
        self.table_result.setColumnCount(len(self.result_df.columns))
        self.table_result.setHorizontalHeaderLabels(self.result_df.columns)
        
        for i in range(len(self.result_df)):
            for j in range(len(self.result_df.columns)):
                value = self.result_df.iloc[i, j]
                item = QTableWidgetItem(str(value))
                
                col_name = self.result_df.columns[j]
                status = self.result_df.iloc[i, len(self.result_df.columns) - 1]
                
                if col_name == "Статус":
                    if status == "РАСХОЖДЕНИЕ":
                        item.setBackground(QBrush(QColor(255, 150, 100)))
                    elif status == "ОК":
                        item.setBackground(QBrush(QColor(150, 255, 150)))
                    else:
                        item.setBackground(QBrush(QColor(255, 200, 100)))
                
                elif col_name == "Разница":
                    try:
                        # Парсим строку с запятой для определения цвета
                        diff_str = str(value).replace(',', '.')
                        diff_val = float(diff_str)
                        if diff_val > 0:
                            item.setForeground(QBrush(QColor(0, 128, 0)))
                        elif diff_val < 0:
                            item.setForeground(QBrush(QColor(255, 0, 0)))
                    except:
                        pass
                
                # Выравнивание чисел по правому краю
                if col_name in ["Значение в файле 1", "Значение в файле 2", "Разница"]:
                    try:
                        test_val = str(value).replace(',', '.')
                        float(test_val)
                        item.setTextAlignment(Qt.AlignRight | Qt.AlignVCenter)
                    except:
                        item.setTextAlignment(Qt.AlignLeft | Qt.AlignVCenter)
                
                self.table_result.setItem(i, j, item)
        
        self.table_result.resizeColumnsToContents()
    
    def save_result(self):
        """Сохранение результата в Excel с сохранением формата чисел (запятая)"""
        file_path, _ = QFileDialog.getSaveFileName(
            self, "Сохранить результат", "",
            "Excel файлы (*.xlsx);;Все файлы (*)"
        )
        
        if file_path:
            try:
                # Создаем копию DataFrame для сохранения с правильным форматированием
                df_to_save = self.result_df.copy()
                
                # Форматируем числовые колонки с запятой
                for col in ["Значение в файле 1", "Значение в файле 2", "Разница"]:
                    if col in df_to_save.columns:
                        df_to_save[col] = df_to_save[col].apply(
                            lambda x: self.format_number_with_comma(x) if x != "ОТСУТСТВУЕТ" and x != "-" else x
                        )
                
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df_to_save.to_excel(writer, index=False, sheet_name='Результаты')
                    
                    from openpyxl.styles import PatternFill, Font, Alignment
                    workbook = writer.book
                    worksheet = writer.sheets['Результаты']
                    
                    green_fill = PatternFill(start_color="96FF96", end_color="96FF96", fill_type="solid")
                    orange_fill = PatternFill(start_color="FF9664", end_color="FF9664", fill_type="solid")
                    yellow_fill = PatternFill(start_color="FFC864", end_color="FFC864", fill_type="solid")
                    
                    green_font = Font(color="008000")
                    red_font = Font(color="FF0000")
                    right_align = Alignment(horizontal="right", vertical="center")
                    
                    status_col = None
                    diff_col = None
                    
                    for col_idx, col_name in enumerate(df_to_save.columns, 1):
                        if col_name == "Статус":
                            status_col = col_idx
                        elif col_name == "Разница":
                            diff_col = col_idx
                    
                    for row_idx in range(2, len(df_to_save) + 2):
                        if status_col:
                            status_cell = worksheet.cell(row_idx, status_col)
                            status = status_cell.value
                            if status == "ОК":
                                status_cell.fill = green_fill
                            elif status == "РАСХОЖДЕНИЕ":
                                status_cell.fill = orange_fill
                            else:
                                status_cell.fill = yellow_fill
                        
                        if diff_col:
                            diff_cell = worksheet.cell(row_idx, diff_col)
                            try:
                                # Парсим строку с запятой
                                diff_val = float(str(diff_cell.value).replace(',', '.'))
                                if diff_val > 0:
                                    diff_cell.font = green_font
                                elif diff_val < 0:
                                    diff_cell.font = red_font
                                diff_cell.alignment = right_align
                            except:
                                diff_cell.alignment = right_align
                
                QMessageBox.information(self, "Успех", f"Результат сохранен в {file_path}")
            except Exception as e:
                QMessageBox.critical(self, "Ошибка", f"Ошибка при сохранении: {str(e)}")


if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelComparator()
    window.show()
    sys.exit(app.exec())
