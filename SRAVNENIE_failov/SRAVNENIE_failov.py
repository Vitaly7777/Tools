"""
Excel Loader с ООП архитектурой
Сравнение двух файлов с умным поиском и сортировкой
"""
from __future__ import annotations

import json
import re
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any
from difflib import SequenceMatcher

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from PySide6 import QtCore, QtGui, QtWidgets

# ============================================================================
# КОНСТАНТЫ
# ============================================================================
DEFAULT_HEADER_SCAN_ROWS = 50
DEFAULT_PREVIEW_PAGE_SIZE = 50
HEADER_LOOKAHEAD_ROWS = 6
HEADER_AUTO_ACCEPT_THRESHOLD = 0.58
HEADER_MIN_SCORE = 6.0
SERVICE_WORD_PATTERNS = (
    "содержание", "описание", "итого", "раздел", "лист", "комментар",
    "summary", "description", "total", "section",
)

TYPE_MAPPING = {
    "auto": None,
    "string": "object",
    "Int64": "Int64",
    "float64": "float64",
    "integer": "int64",
    "boolean": "bool",
}


# ============================================================================
# DATACLASSES
# ============================================================================
@dataclass
class PreviewPage:
    columns: list[str]
    rows: list[list[str]]
    total_rows: int
    offset: int
    limit: int


@dataclass
class AppState:
    """Состояние приложения"""
    file_path: str = ""
    sheet_name: str = ""
    header_row: int = 0
    offset: int = 0
    total_rows: int = 0
    page_size: int = DEFAULT_PREVIEW_PAGE_SIZE
    selected_columns: list[str] = None
    column_types: dict[str, str] = None

    def __post_init__(self):
        if self.selected_columns is None:
            self.selected_columns = []
        if self.column_types is None:
            self.column_types = {}

    def is_file_selected(self) -> bool:
        return bool(self.file_path)

    def is_configured(self) -> bool:
        return bool(self.file_path and self.sheet_name)

    def reset(self) -> None:
        self.offset = 0
        self.total_rows = 0


# ============================================================================
# СРАВНЕНИЕ СТРОК
# ============================================================================
class RowComparator:
    """Класс для сравнения строк с использованием SequenceMatcher"""
    
    @staticmethod
    def calculate_similarity(row1: list[str], row2: list[str]) -> float:
        if not row1 or not row2:
            return 0.0
        str1 = "|".join(str(v).lower().strip() for v in row1)
        str2 = "|".join(str(v).lower().strip() for v in row2)
        matcher = SequenceMatcher(None, str1, str2)
        return matcher.ratio()
    
    @staticmethod
    def find_best_match(row: list[str], candidates: list[list[str]], threshold: float = 0.7) -> tuple[int, float]:
        best_index = -1
        best_similarity = 0.0
        for idx, candidate in enumerate(candidates):
            similarity = RowComparator.calculate_similarity(row, candidate)
            if similarity > best_similarity and similarity >= threshold:
                best_similarity = similarity
                best_index = idx
        return best_index, best_similarity
    
    @staticmethod
    def smart_match(rows1: list[list[str]], rows2: list[list[str]]) -> dict[str, Any]:
        result = {
            "full_matches": [],
            "partial_matches": [],
            "partial_low": [],
            "only_in_file1": [],
            "only_in_file2": [],
        }
        
        matched_indices_2 = set()
        
        for idx1, row1 in enumerate(rows1):
            found_match = False
            
            for idx2, row2 in enumerate(rows2):
                if idx2 in matched_indices_2:
                    continue
                if row1 == row2:
                    result["full_matches"].append({
                        "row1": row1, "row2": row2,
                        "similarity": 1.0, "index1": idx1, "index2": idx2,
                    })
                    matched_indices_2.add(idx2)
                    found_match = True
                    break
            
            if not found_match:
                remaining_candidates = [
                    (idx, rows2[idx]) 
                    for idx in range(len(rows2)) 
                    if idx not in matched_indices_2
                ]
                
                if remaining_candidates:
                    best_idx, similarity = RowComparator.find_best_match(
                        row1, [r[1] for r in remaining_candidates], threshold=0.4
                    )
                    
                    if best_idx >= 0:
                        actual_idx2 = remaining_candidates[best_idx][0]
                        if similarity >= 0.7:
                            result["partial_matches"].append({
                                "row1": row1, "row2": rows2[actual_idx2],
                                "similarity": similarity, "index1": idx1, "index2": actual_idx2,
                            })
                        else:
                            result["partial_low"].append({
                                "row1": row1, "row2": rows2[actual_idx2],
                                "similarity": similarity, "index1": idx1, "index2": actual_idx2,
                            })
                        matched_indices_2.add(actual_idx2)
                    else:
                        result["only_in_file1"].append({"row1": row1, "index1": idx1})
                else:
                    result["only_in_file1"].append({"row1": row1, "index1": idx1})
        
        for idx2, row2 in enumerate(rows2):
            if idx2 not in matched_indices_2:
                result["only_in_file2"].append({"row2": row2, "index2": idx2})
        
        return result


# ============================================================================
# EXCEL CORE
# ============================================================================
class ExcelReader:
    """Чтение Excel файлов"""

    def __init__(self, file_path: str):
        self.file_path = self._normalize_path(file_path)

    def _normalize_path(self, file_path: str) -> Path:
        path = Path(str(file_path or "").strip())
        if not str(path):
            raise FileNotFoundError("Не выбран Excel-файл.")
        if not path.exists():
            raise FileNotFoundError(f"Файл не найден: {path}")
        return path

    def get_sheet_names(self) -> list[str]:
        if self.file_path.suffix.lower() == ".csv":
            return [self.file_path.name]
        workbook = pd.ExcelFile(self.file_path)
        return list(workbook.sheet_names)

    def read_raw_frame(self, sheet_name: str, nrows: int) -> pd.DataFrame:
        if self.file_path.suffix.lower() == ".csv":
            return pd.read_csv(self.file_path, header=None, nrows=nrows, dtype=object)
        return pd.read_excel(
            self.file_path, header=None, nrows=nrows, dtype=object,
            sheet_name=str(sheet_name or "").strip() or 0
        )

    def read_dataframe(
        self, sheet_name: str, header_row: int, nrows: int | None,
        usecols: list[str] | None = None, dtype: dict[str, str] | None = None
    ) -> pd.DataFrame:
        header_index = max(0, int(header_row))
        
        if self.file_path.suffix.lower() == ".csv":
            df = pd.read_csv(
                self.file_path, header=header_index, dtype=object, nrows=nrows,
                usecols=usecols if usecols else None
            )
        else:
            df = pd.read_excel(
                self.file_path, header=header_index, dtype=object,
                sheet_name=str(sheet_name or "").strip() or 0, nrows=nrows,
                usecols=usecols if usecols else None
            )
        
        if dtype:
            for col, type_name in dtype.items():
                if col in df.columns:
                    try:
                        if type_name == "integer":
                            # ОКРУГЛЕНИЕ ДО ЦЕЛОГО И ТИП Int64 (для поддержки NaN)
                            df[col] = pd.to_numeric(df[col], errors='coerce').round().astype("Int64")
                        else:
                            mapped_type = TYPE_MAPPING.get(type_name)
                            if mapped_type:
                                df[col] = df[col].astype(mapped_type)
                    except Exception as e:
                        print(f"Ошибка при преобразовании {col} в {type_name}: {e}")
        
        return df

    def get_total_rows(self, sheet_name: str, header_row: int, usecols: list[str] | None = None) -> int:
        frame = self.read_dataframe(sheet_name, header_row, nrows=None, usecols=usecols)
        return int(len(frame.index))

    def get_preview_page(
        self, sheet_name: str, header_row: int, offset: int = 0,
        limit: int = DEFAULT_PREVIEW_PAGE_SIZE, usecols: list[str] | None = None,
        dtype: dict[str, str] | None = None
    ) -> PreviewPage:
        frame = self.read_dataframe(sheet_name, header_row, nrows=None, usecols=usecols, dtype=dtype)
        start = max(0, int(offset))
        page_size = max(1, int(limit))
        end = start + page_size
        page_frame = frame.iloc[start:end]
        
        columns = [str(column) for column in page_frame.columns]
        rows = [
            ["" if pd.isna(value) else str(value) for value in row]
            for row in page_frame.itertuples(index=False, name=None)
        ]
        return PreviewPage(
            columns=columns, rows=rows, total_rows=int(len(frame.index)),
            offset=start, limit=page_size
        )

    def get_all_dataframe(
        self, sheet_name: str, header_row: int,
        usecols: list[str] | None = None, dtype: dict[str, str] | None = None
    ) -> pd.DataFrame:
        return self.read_dataframe(sheet_name, header_row, nrows=None, usecols=usecols, dtype=dtype)


# ============================================================================
# HEADER DETECTION
# ============================================================================
class HeaderDetector:
    """Автоматический поиск строки шапки"""

    def __init__(self, excel_reader: ExcelReader):
        self.reader = excel_reader

    def detect(self, file_path: str, sheet_name: str, scan_rows: int = DEFAULT_HEADER_SCAN_ROWS) -> dict[str, Any]:
        raw_frame = self.reader.read_raw_frame(sheet_name, max(5, int(scan_rows)))
        return self._detect_header_row(raw_frame)

    def build_preview(
        self, sheet_name: str, header_row: int | None = None,
        selected_columns: list[str] | None = None, column_types: dict[str, str] | None = None,
        scan_rows: int = DEFAULT_HEADER_SCAN_ROWS
    ) -> dict[str, Any]:
        raw_frame = self.reader.read_raw_frame(sheet_name, max(5, int(scan_rows)))

        if header_row is None:
            detection = self._detect_header_row(raw_frame)
            detected_header = int(detection["header_row"])
            header_confidence = float(detection["confidence"])
            header_confirmed = bool(detection["auto_accept"])
        else:
            detected_header = max(0, int(header_row))
            header_confidence = 1.0
            header_confirmed = True

        sample_frame = self.reader.read_dataframe(sheet_name, detected_header, nrows=50)

        raw_rows = [
            ["" if pd.isna(value) else str(value) for value in raw_frame.iloc[row_index].tolist()]
            for row_index in range(len(raw_frame.index))
        ]

        selected = set(selected_columns or [])
        normalized_column_types = dict(column_types or {})
        rows: list[dict[str, Any]] = []
        for column_name in sample_frame.columns:
            normalized_name = str(column_name).strip()
            if not normalized_name:
                continue
            rows.append({
                "name": normalized_name,
                "detected_type": self._detect_type(sample_frame[column_name]),
                "enabled": True if not selected else normalized_name in selected,
                "read_as": normalized_column_types.get(normalized_name, "auto"),
            })

        return {
            "header_row": detected_header,
            "header_confidence": header_confidence,
            "header_confirmed": header_confirmed,
            "raw_rows": raw_rows,
            "rows": rows,
        }

    def _detect_header_row(self, frame: pd.DataFrame) -> dict[str, Any]:
        if frame.empty:
            return {
                "header_row": 0, "confidence": 0.0, "auto_accept": False,
                "explanation": "Нет строк для анализа.", "candidates": [],
            }

        candidates = []
        for row_index in range(len(frame.index)):
            features = self._extract_row_features(frame, row_index)
            header_score, header_reasons = self._score_header_likeness(features)
            data_score, metrics, data_reasons = self._score_data_below(frame, row_index, features)
            penalty, anti_reasons = self._score_anti_patterns(features)
            final_score = header_score + data_score - penalty
            candidates.append({
                "row_index": row_index,
                "final_score": round(final_score, 3),
                "header_likeness": round(header_score, 3),
                "metrics": metrics,
            })

        ordered = sorted(candidates, key=lambda item: (item["final_score"], -item["row_index"]), reverse=True)
        best = ordered[0]
        second_score = ordered[1]["final_score"] if len(ordered) > 1 else best["final_score"] - 5.0
        score_gap = float(best["final_score"] - second_score)
        strength = max(0.0, min(1.0, best["final_score"] / 18.0))
        separation = max(0.0, min(1.0, score_gap / 6.0))
        confidence = round((strength * 0.6) + (separation * 0.4), 3)
        auto_accept = (
            best["final_score"] >= HEADER_MIN_SCORE and
            confidence >= HEADER_AUTO_ACCEPT_THRESHOLD
        )

        return {
            "header_row": int(best["row_index"]),
            "confidence": confidence,
            "auto_accept": auto_accept,
            "explanation": "Шапка найдена" if auto_accept else "Шапка найдена с низкой уверенностью",
            "candidates": ordered[:5],
        }

    def _extract_row_features(self, frame: pd.DataFrame, row_index: int) -> dict[str, Any]:
        row = frame.iloc[row_index].tolist()
        positions, labels = [], []
        text_like_count = short_label_count = 0

        for column_index, value in enumerate(row):
            text = self._normalize_cell_text(value)
            if not text:
                continue
            positions.append(column_index)
            labels.append(text)
            if any(char.isalpha() for char in text):
                text_like_count += 1
            if len(text) <= 32 and len(text.split()) <= 5:
                short_label_count += 1

        unique_count = len({label.casefold() for label in labels})
        return {
            "filled_count": len(labels),
            "unique_count": unique_count,
            "duplicate_count": len(labels) - unique_count,
            "text_like_count": text_like_count,
            "short_label_count": short_label_count,
        }

    def _score_header_likeness(self, features: dict[str, Any]) -> tuple[float, list[str]]:
        if features["filled_count"] == 0:
            return -10.0, ["пустая строка"]
        score = features["filled_count"] * 1.4 + features["text_like_count"] * 1.2 + features["short_label_count"] * 0.8
        score -= features["duplicate_count"] * 3.5
        reasons = []
        if features["filled_count"] >= 3:
            reasons.append("много заполненных ячеек")
        if features["text_like_count"] >= max(1, features["filled_count"] // 2):
            reasons.append("заголовки выглядят текстовыми")
        return score, reasons

    def _score_data_below(self, frame: pd.DataFrame, row_index: int, features: dict[str, Any]) -> tuple[float, dict[str, Any], list[str]]:
        score = 0.0
        metrics = {"stable_rows": 0, "filled_rows": 0, "typed_value_hits": 0}
        reasons = []
        if row_index + 1 < len(frame.index):
            score = 2.0
            metrics["stable_rows"] = 1
            reasons.append("ниже есть данные")
        else:
            score = -5.0
            reasons.append("ниже нет данных")
        return score, metrics, reasons

    def _score_anti_patterns(self, features: dict[str, Any]) -> tuple[float, list[str]]:
        penalty = features["duplicate_count"] * 1.6
        reasons = []
        if features["duplicate_count"]:
            reasons.append("дубли заголовков")
        return penalty, reasons

    def _normalize_cell_text(self, value: Any) -> str:
        text = "" if pd.isna(value) else str(value).strip()
        return "" if text.lower() == "nan" else text

    def _detect_type(self, series: pd.Series) -> str:
        cleaned = series.dropna()
        if cleaned.empty:
            return "unknown"
        if pd.api.types.is_integer_dtype(cleaned):
            return "integer"
        if pd.api.types.is_float_dtype(cleaned):
            return "float"
        return "string"


# ============================================================================
# UI КОМПОНЕНТЫ
# ============================================================================
class ColumnsTable(QtWidgets.QTableWidget):
    """Таблица выбора колонок"""
    
    stateChanged = QtCore.Signal()
    TYPE_OPTIONS = [("auto", "auto"), ("string", "string"), ("Int64", "Int64"), ("float64", "float64"), ("integer", "integer")]

    def __init__(self, parent=None):
        super().__init__(0, 4, parent)
        self.setHorizontalHeaderLabels(["Колонка", "Определено", "Читать", "Тип"])
        self.verticalHeader().setVisible(False)
        self.horizontalHeader().setSectionResizeMode(0, QtWidgets.QHeaderView.ResizeMode.Stretch)
        
        # ДОБАВЛЕНО: Клик по заголовку для "Выбрать/Снять все"
        self.horizontalHeader().sectionClicked.connect(self._on_header_clicked)
        self._all_selected = True

    def _on_header_clicked(self, logical_index):
        if logical_index == 2:  # Колонка "Читать"
            self._all_selected = not self._all_selected
            for row in range(self.rowCount()):
                wrapper = self.cellWidget(row, 2)
                if wrapper:
                    checkbox = wrapper.findChild(QtWidgets.QCheckBox)
                    if checkbox:
                        checkbox.setChecked(self._all_selected)
            self.stateChanged.emit()

    def load_columns(self, rows: list[dict[str, Any]]) -> None:
        self.setRowCount(0)
        for row_index, row in enumerate(rows):
            self.insertRow(row_index)
            self.setItem(row_index, 0, QtWidgets.QTableWidgetItem(str(row["name"])))
            self.setItem(row_index, 1, QtWidgets.QTableWidgetItem(str(row["detected_type"])))
            
            checkbox = QtWidgets.QCheckBox()
            checkbox.setChecked(True)
            checkbox.stateChanged.connect(self.stateChanged.emit)
            wrapper = QtWidgets.QWidget()
            wrapper_layout = QtWidgets.QHBoxLayout(wrapper)
            wrapper_layout.setContentsMargins(0, 0, 0, 0)
            wrapper_layout.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
            wrapper_layout.addWidget(checkbox)
            self.setCellWidget(row_index, 2, wrapper)

            combo = QtWidgets.QComboBox()
            for value, label in self.TYPE_OPTIONS:
                combo.addItem(label, value)
            combo.currentIndexChanged.connect(self.stateChanged.emit)
            self.setCellWidget(row_index, 3, combo)

    def get_selected_columns(self) -> list[str]:
        selected = []
        for row in range(self.rowCount()):
            name_item = self.item(row, 0)
            wrapper = self.cellWidget(row, 2)
            if wrapper and (checkbox := wrapper.findChild(QtWidgets.QCheckBox)):
                if checkbox.isChecked():
                    selected.append(name_item.text())
        return selected

    def get_column_types(self) -> dict[str, str]:
        types_dict = {}
        for row in range(self.rowCount()):
            name_item = self.item(row, 0)
            wrapper = self.cellWidget(row, 2)
            combo = self.cellWidget(row, 3)
            if wrapper and (checkbox := wrapper.findChild(QtWidgets.QCheckBox)):
                if checkbox.isChecked():
                    column_name = name_item.text()
                    if combo and isinstance(combo, QtWidgets.QComboBox):
                        type_value = str(combo.currentData() or "auto")
                        if type_value != "auto":
                            types_dict[column_name] = type_value
        return types_dict

    def has_selected_columns(self) -> bool:
        return len(self.get_selected_columns()) > 0


class DataComparisonTable(QtWidgets.QTableWidget):
    """Таблица сравнения двух файлов с умным поиском"""

    def __init__(self, parent=None):
        super().__init__(parent)
        self.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.verticalHeader().setVisible(False)
        self.setAlternatingRowColors(True)
        self.comparison_results = []
        self.table_col_count_file1 = 0
        self.table_col_count_file2 = 0

    def load_comparison_smart(self, rows1: list[list[str]], rows2: list[list[str]], 
                             columns1: list[str], columns2: list[str]) -> None:
        """Загрузить умное сравнение"""
        comparison_result = RowComparator.smart_match(rows1, rows2)
        
        self.clear()
        self.setRowCount(0)
        
        headers = columns1 + ["📊 Сравнение"] + columns2
        self.setColumnCount(len(headers))
        self.setHorizontalHeaderLabels(headers)
        
        self.table_col_count_file1 = len(columns1)
        self.table_col_count_file2 = len(columns2)
        self.comparison_results = []
        
        row_num = 0
        
        if comparison_result["full_matches"]:
            row_num = self._add_section_header(row_num, "✓ ПОЛНЫЕ СОВПАДЕНИЯ")
            for match in comparison_result["full_matches"]:
                row_num = self._add_comparison_row(
                    row_num, match["row1"], match["row2"],
                    columns1, columns2, "full_match", "✓ СОВПАДАЕТ"
                )
        
        if comparison_result["partial_matches"]:
            row_num = self._add_section_header(row_num, "⚡ ЧАСТИЧНЫЕ СОВПАДЕНИЯ (70-99%)")
            for match in sorted(comparison_result["partial_matches"], 
                              key=lambda x: x["similarity"], reverse=True):
                percent = int(match["similarity"] * 100)
                row_num = self._add_comparison_row(
                    row_num, match["row1"], match["row2"],
                    columns1, columns2, "partial_match", f"⚡ {percent}% совпадение"
                )
        
        if comparison_result["partial_low"]:
            row_num = self._add_section_header(row_num, "⚠️ НИЗКИЕ СОВПАДЕНИЯ (40-70%)")
            for match in sorted(comparison_result["partial_low"],
                              key=lambda x: x["similarity"], reverse=True):
                percent = int(match["similarity"] * 100)
                row_num = self._add_comparison_row(
                    row_num, match["row1"], match["row2"],
                    columns1, columns2, "partial_low", f"⚠️ {percent}% совпадение"
                )
        
        if comparison_result["only_in_file1"]:
            row_num = self._add_section_header(row_num, "❌ ТОЛЬКО В ФАЙЛЕ 1")
            for match in comparison_result["only_in_file1"]:
                row_num = self._add_comparison_row(
                    row_num, match["row1"], [""] * len(columns2),
                    columns1, columns2, "only_file1", "❌ Нет в файле 2"
                )
        
        if comparison_result["only_in_file2"]:
            row_num = self._add_section_header(row_num, "❌ ТОЛЬКО В ФАЙЛЕ 2")
            for match in comparison_result["only_in_file2"]:
                row_num = self._add_comparison_row(
                    row_num, [""] * len(columns1), match["row2"],
                    columns1, columns2, "only_file2", "❌ Нет в файле 1"
                )
        
        self.resizeColumnsToContents()
        self.viewport().update()

    def _add_section_header(self, row_num: int, title: str) -> int:
        """Добавить заголовок секции"""
        self.insertRow(row_num)
        total_cols = self.table_col_count_file1 + 1 + self.table_col_count_file2
        for col in range(total_cols):
            item = QtWidgets.QTableWidgetItem(title if col == 0 else "")
            item.setBackground(QtGui.QColor(80, 80, 80))
            item.setForeground(QtGui.QColor(255, 255, 255))
            font = item.font()
            font.setBold(True)
            font.setPointSize(11)
            item.setFont(font)
            item.setTextAlignment(int(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter))
            self.setItem(row_num, col, item)
        self.setRowHeight(row_num, 25)
        self.comparison_results.append("section")
        return row_num + 1

    def _add_comparison_row(self, row_num: int, row1: list[str], row2: list[str],
                           columns1: list[str], columns2: list[str],
                           match_type: str, comparison_text: str) -> int:
        """Добавить строку сравнения"""
        self.insertRow(row_num)
        col_num = 0
        
        # Данные файла 1
        for col_index in range(len(columns1)):
            value = row1[col_index] if col_index < len(row1) else ""
            item = QtWidgets.QTableWidgetItem(str(value))
            item.setTextAlignment(int(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter))
            self.setItem(row_num, col_num, item)
            col_num += 1
        
        # Столбец сравнения
        comparison_item = QtWidgets.QTableWidgetItem(comparison_text)
        comparison_item.setTextAlignment(int(QtCore.Qt.AlignmentFlag.AlignCenter | QtCore.Qt.AlignmentFlag.AlignVCenter))
        comparison_item.setFont(QtGui.QFont("Arial", 9, QtGui.QFont.Weight.Bold))
        
        if match_type == "full_match":
            comparison_item.setBackground(QtGui.QColor(0, 200, 0))
            comparison_item.setForeground(QtGui.QColor(255, 255, 255))
        elif match_type == "partial_match":
            comparison_item.setBackground(QtGui.QColor(255, 165, 0))
            comparison_item.setForeground(QtGui.QColor(0, 0, 0))
        elif match_type == "partial_low":
            comparison_item.setBackground(QtGui.QColor(255, 255, 0))
            comparison_item.setForeground(QtGui.QColor(0, 0, 0))
        elif match_type == "only_file1":
            comparison_item.setBackground(QtGui.QColor(255, 80, 80))
            comparison_item.setForeground(QtGui.QColor(255, 255, 255))
        elif match_type == "only_file2":
            comparison_item.setBackground(QtGui.QColor(80, 180, 255))
            comparison_item.setForeground(QtGui.QColor(0, 0, 0))
        
        self.setItem(row_num, col_num, comparison_item)
        col_num += 1
        
        # Данные файла 2
        for col_index in range(len(columns2)):
            value = row2[col_index] if col_index < len(row2) else ""
            item = QtWidgets.QTableWidgetItem(str(value))
            item.setTextAlignment(int(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter))
            self.setItem(row_num, col_num, item)
            col_num += 1
        
        self.comparison_results.append(match_type)
        return row_num + 1

    def export_to_excel(self, file_path: str, file1_name: str, file2_name: str) -> None:
        """Экспортировать в Excel с форматированием"""
        try:
            wb = Workbook()
            ws = wb.active
            ws.title = "Сравнение"

            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            match_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            partial_fill = PatternFill(start_color="FFD966", end_color="FFD966", fill_type="solid")
            partial_low_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            file1_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
            file2_fill = PatternFill(start_color="4DA6FF", end_color="4DA6FF", fill_type="solid")
            section_fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            total_cols = self.table_col_count_file1 + 1 + self.table_col_count_file2
            
            col_num = 1
            for col_index in range(self.table_col_count_file1):
                cell = ws.cell(row=1, column=col_num)
                header_item = self.horizontalHeaderItem(col_index)
                cell.value = f"{file1_name}: {header_item.text() if header_item else ''}"
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                col_num += 1
            
            cell = ws.cell(row=1, column=col_num)
            cell.value = "Сравнение"
            cell.fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF", size=11)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            col_num += 1
            
            for col_index in range(self.table_col_count_file2):
                cell = ws.cell(row=1, column=col_num)
                header_item = self.horizontalHeaderItem(self.table_col_count_file1 + 1 + col_index)
                cell.value = f"{file2_name}: {header_item.text() if header_item else ''}"
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border
                col_num += 1
            
            for row_index in range(self.rowCount()):
                match_type = self.comparison_results[row_index] if row_index < len(self.comparison_results) else "unknown"
                
                for col_index in range(total_cols):
                    item = self.item(row_index, col_index)
                    cell = ws.cell(row=row_index + 2, column=col_index + 1)
                    cell.value = item.text() if item else ""
                    cell.border = border
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    
                    if col_index == self.table_col_count_file1:
                        cell.font = Font(bold=True)
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        if match_type == "full_match":
                            cell.fill = match_fill
                        elif match_type == "partial_match":
                            cell.fill = partial_fill
                        elif match_type == "partial_low":
                            cell.fill = partial_low_fill
                        elif match_type == "only_file1":
                            cell.fill = file1_fill
                            cell.font = Font(bold=True, color="FFFFFF")
                        elif match_type == "only_file2":
                            cell.fill = file2_fill
                            cell.font = Font(bold=True, color="FFFFFF")
                        elif match_type == "section":
                            cell.fill = section_fill
                            cell.font = Font(bold=True, color="FFFFFF")
                    
                    if match_type == "section":
                        cell.fill = section_fill
                        cell.font = Font(bold=True, color="FFFFFF")
            
            for col in range(1, total_cols + 1):
                ws.column_dimensions[get_column_letter(col)].width = 25
            ws.row_dimensions[1].height = 30
            
            wb.save(file_path)
        except Exception as e:
            raise


class FileConfigPanel(QtWidgets.QWidget):
    """Панель конфигурации одного файла"""

    configured = QtCore.Signal()

    def __init__(self, file_number: int, parent=None):
        super().__init__(parent)
        self.file_number = file_number
        self.reader: ExcelReader | None = None
        self.detector: HeaderDetector | None = None
        self.state = AppState()
        self._is_loading = False
        self._build_ui()

    def _build_ui(self) -> None:
        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        title = QtWidgets.QLabel(f"📁 Файл #{self.file_number}")
        title_font = title.font()
        title_font.setBold(True)
        title_font.setPointSize(12)
        title.setFont(title_font)
        layout.addWidget(title)

        file_layout = QtWidgets.QHBoxLayout()
        self.file_edit = QtWidgets.QLineEdit()
        self.file_edit.setReadOnly(True)
        self.choose_btn = QtWidgets.QPushButton("Выбрать файл")
        self.choose_btn.clicked.connect(self._choose_file)
        file_layout.addWidget(QtWidgets.QLabel("Файл:"))
        file_layout.addWidget(self.file_edit, 1)
        file_layout.addWidget(self.choose_btn)
        layout.addLayout(file_layout)

        info = QtWidgets.QHBoxLayout()
        self.sheet_label = QtWidgets.QLabel("Лист: -")
        self.header_label = QtWidgets.QLabel("Шапка: -")
        self.columns_label = QtWidgets.QLabel("Колонок: -")
        info.addWidget(self.sheet_label)
        info.addSpacing(15)
        info.addWidget(self.header_label)
        info.addSpacing(15)
        info.addWidget(self.columns_label)
        info.addStretch(1)
        layout.addLayout(info)

        self.configure_btn = QtWidgets.QPushButton("Настроить этот файл")
        self.configure_btn.setEnabled(False)
        self.configure_btn.clicked.connect(self._configure)
        layout.addWidget(self.configure_btn)

        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.Shape.HLine)
        layout.addWidget(separator)

    def _choose_file(self) -> None:
        file_path, _ = QtWidgets.QFileDialog.getOpenFileName(
            self, f"Выбрать файл #{self.file_number}", "",
            "Excel (*.xlsx *.xls *.csv);;Все (*)"
        )
        if not file_path:
            return
        self.state = AppState(file_path=file_path)
        self.reader = ExcelReader(file_path)
        self.detector = HeaderDetector(self.reader)
        self.file_edit.setText(file_path)
        self.sheet_label.setText("Лист: -")
        self.header_label.setText("Шапка: -")
        self.columns_label.setText("Колонок: -")
        self.configure_btn.setEnabled(True)

    def _configure(self) -> None:
        if not self.state.file_path:
            return
        dialog = ExcelWizardDialog(self.state.file_path, self)
        if dialog.exec() != dialog.DialogCode.Accepted:
            return
        result = dialog.get_result()
        self.state.sheet_name = result["sheet_name"]
        self.state.header_row = result["header"]
        self.state.selected_columns = result["selected_columns"]
        self.state.column_types = result["column_types"]
        self.state.offset = 0
        self.state.total_rows = self.reader.get_total_rows(
            self.state.sheet_name, self.state.header_row,
            usecols=self.state.selected_columns if self.state.selected_columns else None
        )
        self.sheet_label.setText(f"Лист: {self.state.sheet_name}")
        self.header_label.setText(f"Шапка: строка {self.state.header_row + 1}")
        self.columns_label.setText(f"Колонок: {len(self.state.selected_columns)}")
        self.configured.emit()

    def get_all_rows(self) -> list[list[str]]:
        if not self.state.is_configured():
            raise ValueError(f"Файл #{self.file_number} не настроен")
        df = self.reader.get_all_dataframe(
            self.state.sheet_name, self.state.header_row,
            usecols=self.state.selected_columns if self.state.selected_columns else None,
            dtype=self.state.column_types if self.state.column_types else None
        )
        rows = [
            ["" if pd.isna(value) else str(value) for value in row]
            for row in df.itertuples(index=False, name=None)
        ]
        return rows

    def get_columns(self) -> list[str]:
        return self.state.selected_columns

    def get_file_name(self) -> str:
        if self.state.file_path:
            return Path(self.state.file_path).stem
        return f"File{self.file_number}"


# ============================================================================
# WIZARD
# ============================================================================
class ExcelWizardDialog(QtWidgets.QDialog):
    """Диалог настройки Excel"""

    def __init__(self, file_path: str, parent=None):
        super().__init__(parent)
        self.reader = ExcelReader(file_path)
        self.detector = HeaderDetector(self.reader)
        self.state = AppState(file_path=file_path)
        self._setup_ui()
        self._load_sheets()

    def _setup_ui(self) -> None:
        self.setWindowTitle("Настройка Excel")
        self.resize(900, 700)
        layout = QtWidgets.QVBoxLayout(self)

        self.steps_label = QtWidgets.QLabel("Шаг 1 из 3: Выбор листа")
        layout.addWidget(self.steps_label)

        self.stack = QtWidgets.QStackedWidget()
        layout.addWidget(self.stack, 1)

        page1 = QtWidgets.QWidget()
        page1_layout = QtWidgets.QVBoxLayout(page1)
        self.sheet_list = QtWidgets.QListWidget()
        self.sheet_list.itemSelectionChanged.connect(self._on_sheet_selected)
        page1_layout.addWidget(self.sheet_list)
        self.stack.addWidget(page1)

        page2 = QtWidgets.QWidget()
        page2_layout = QtWidgets.QVBoxLayout(page2)
        controls = QtWidgets.QHBoxLayout()
        self.header_mode = QtWidgets.QComboBox()
        self.header_mode.addItem("Авто", "auto")
        self.header_mode.addItem("Вручную", "manual")
        self.header_mode.currentIndexChanged.connect(self._on_header_mode_changed)
        self.header_row_spin = QtWidgets.QSpinBox()
        self.header_row_spin.setRange(1, 10000)
        self.header_row_spin.valueChanged.connect(self._on_manual_header_changed)
        controls.addWidget(QtWidgets.QLabel("Режим:"))
        controls.addWidget(self.header_mode)
        controls.addWidget(QtWidgets.QLabel("Строка:"))
        controls.addWidget(self.header_row_spin)
        controls.addStretch(1)
        page2_layout.addLayout(controls)
        self.header_table = QtWidgets.QTableWidget()
        self.header_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        page2_layout.addWidget(self.header_table, 1)
        self.stack.addWidget(page2)

        page3 = QtWidgets.QWidget()
        page3_layout = QtWidgets.QVBoxLayout(page3)
        info_label = QtWidgets.QLabel("Выберите колонки, которые нужно читать.\nВыберите тип данных для каждой колонки:")
        info_label.setWordWrap(True)
        page3_layout.addWidget(info_label)
        self.columns_table = ColumnsTable()
        self.columns_table.stateChanged.connect(self._on_columns_changed)
        page3_layout.addWidget(self.columns_table)
        self.stack.addWidget(page3)

        buttons = QtWidgets.QHBoxLayout()
        self.back_btn = QtWidgets.QPushButton("Назад")
        self.next_btn = QtWidgets.QPushButton("Далее")
        self.save_btn = QtWidgets.QPushButton("Сохранить")
        self.cancel_btn = QtWidgets.QPushButton("Отмена")
        self.back_btn.clicked.connect(self._go_back)
        self.next_btn.clicked.connect(self._go_next)
        self.save_btn.clicked.connect(self.accept)
        self.cancel_btn.clicked.connect(self.reject)
        buttons.addWidget(self.back_btn)
        buttons.addWidget(self.next_btn)
        buttons.addStretch(1)
        buttons.addWidget(self.save_btn)
        buttons.addWidget(self.cancel_btn)
        layout.addLayout(buttons)
        self._update_buttons()

    def _load_sheets(self) -> None:
        for sheet_name in self.reader.get_sheet_names():
            self.sheet_list.addItem(sheet_name)
        if self.sheet_list.count() > 0:
            self.sheet_list.setCurrentRow(0)

    def _on_sheet_selected(self) -> None:
        item = self.sheet_list.currentItem()
        if item:
            self.state.sheet_name = item.text()
            self._refresh_header()

    def _on_header_mode_changed(self) -> None:
        if self.header_mode.currentData() == "manual":
            self._on_manual_header_changed()
        else:
            self._refresh_header()

    def _on_manual_header_changed(self) -> None:
        if self.header_mode.currentData() != "manual":
            return
        if not self.state.sheet_name:
            return
        try:
            manual_header_row = max(0, self.header_row_spin.value() - 1)
            raw_frame = self.reader.read_raw_frame(self.state.sheet_name, nrows=manual_header_row + 20)
            raw_rows = [
                ["" if pd.isna(value) else str(value) for value in raw_frame.iloc[row_index].tolist()]
                for row_index in range(len(raw_frame.index))
            ]
            self._fill_header_table(raw_rows, manual_header_row)
            self.state.header_row = manual_header_row
            preview = self.detector.build_preview(self.state.sheet_name, header_row=manual_header_row)
            self.columns_table.load_columns(preview["rows"])
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Ошибка", f"Не могу прочитать строку: {str(e)}")

    def _refresh_header(self) -> None:
        if not self.state.sheet_name:
            return
        try:
            preview = self.detector.build_preview(self.state.sheet_name)
            header_row = preview["header_row"]
            self.header_row_spin.blockSignals(True)
            self.header_row_spin.setValue(header_row + 1)
            self.header_row_spin.blockSignals(False)
            self.state.header_row = header_row
            self._fill_header_table(preview["raw_rows"], header_row)
            self.columns_table.load_columns(preview["rows"])
        except Exception as e:
            QtWidgets.QMessageBox.warning(self, "Ошибка", str(e))

    def _fill_header_table(self, raw_rows: list[list[str]], header_row: int) -> None:
        if not raw_rows:
            self.header_table.setRowCount(0)
            return
        col_count = max(len(row) for row in raw_rows) if raw_rows else 0
        self.header_table.setRowCount(len(raw_rows))
        self.header_table.setColumnCount(col_count)
        row_labels = [str(i + 1) for i in range(len(raw_rows))]
        self.header_table.setVerticalHeaderLabels(row_labels)
        col_labels = [chr(65 + i) if i < 26 else f"A{chr(65 + i - 26)}" for i in range(col_count)]
        self.header_table.setHorizontalHeaderLabels(col_labels)
        for row_idx, row in enumerate(raw_rows):
            for col_idx in range(col_count):
                value = row[col_idx] if col_idx < len(row) else ""
                item = QtWidgets.QTableWidgetItem(value)
                if row_idx == header_row:
                    item.setBackground(QtGui.QColor(255, 255, 0))
                    font = item.font()
                    font.setBold(True)
                    item.setFont(font)
                self.header_table.setItem(row_idx, col_idx, item)
        self.header_table.resizeColumnsToContents()

    def _on_columns_changed(self) -> None:
        self.state.selected_columns = self.columns_table.get_selected_columns()
        self.state.column_types = self.columns_table.get_column_types()

    def _go_back(self) -> None:
        self.stack.setCurrentIndex(max(0, self.stack.currentIndex() - 1))
        self._update_buttons()

    def _go_next(self) -> None:
        if self.stack.currentIndex() == 2 and not self.columns_table.has_selected_columns():
            QtWidgets.QMessageBox.warning(self, "Ошибка", "Выберите хотя бы одну колонку!")
            return
        self.stack.setCurrentIndex(min(2, self.stack.currentIndex() + 1))
        self._update_buttons()

    def _update_buttons(self) -> None:
        index = self.stack.currentIndex()
        self.steps_label.setText(f"Шаг {index + 1} из 3: {['Листы', 'Шапка', 'Колонки'][index]}")
        self.back_btn.setEnabled(index > 0)
        self.next_btn.setVisible(index < 2)
        self.save_btn.setVisible(index == 2)

    def get_result(self) -> dict[str, Any]:
        return {
            "file_path": self.state.file_path,
            "sheet_name": self.state.sheet_name,
            "header": self.state.header_row,
            "selected_columns": self.columns_table.get_selected_columns(),
            "column_types": self.columns_table.get_column_types(),
        }
# ============================================================================
# MAIN WINDOW
# ============================================================================
class MainWindow(QtWidgets.QMainWindow):
    """Главное окно приложения"""

    def __init__(self):
        super().__init__()
        self.setWindowTitle("Excel Comparator - Умное сравнение двух файлов")
        self.resize(1600, 900)
        self._build_ui()

    def _build_ui(self) -> None:
        """Построить UI"""
        central = QtWidgets.QWidget()
        self.setCentralWidget(central)
        layout = QtWidgets.QVBoxLayout(central)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(10)

        # Верхняя часть - конфигурация файлов
        config_layout = QtWidgets.QHBoxLayout()
        
        self.config_panel_1 = FileConfigPanel(1)
        self.config_panel_1.configured.connect(self._on_file_configured)
        config_layout.addWidget(self.config_panel_1, 1)
        
        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.Shape.VLine)
        config_layout.addWidget(separator)
        
        self.config_panel_2 = FileConfigPanel(2)
        self.config_panel_2.configured.connect(self._on_file_configured)
        config_layout.addWidget(self.config_panel_2, 1)
        
        layout.addLayout(config_layout, 0)

        main_separator = QtWidgets.QFrame()
        main_separator.setFrameShape(QtWidgets.QFrame.Shape.HLine)
        main_separator.setLineWidth(2)
        layout.addWidget(main_separator)

        # Нижняя часть - результаты сравнения
        bottom_layout = QtWidgets.QVBoxLayout()
        
        self.info_label = QtWidgets.QLabel("Настройте оба файла для начала сравнения")
        bottom_layout.addWidget(self.info_label)
        
        controls = QtWidgets.QHBoxLayout()
        self.refresh_btn = QtWidgets.QPushButton("🔄 Обновить сравнение")
        self.refresh_btn.setEnabled(False)
        self.refresh_btn.clicked.connect(self._load_comparison)
        self.export_btn = QtWidgets.QPushButton("💾 Экспортировать отчет (.xlsx)")
        self.export_btn.setEnabled(False)
        self.export_btn.clicked.connect(self._export_report)
        controls.addWidget(self.refresh_btn)
        controls.addWidget(self.export_btn)
        controls.addStretch(1)
        bottom_layout.addLayout(controls)
        
        self.comparison_table = DataComparisonTable()
        bottom_layout.addWidget(self.comparison_table, 1)
        
        layout.addLayout(bottom_layout, 1)

    def _on_file_configured(self) -> None:
        """Файл настроен"""
        if (self.config_panel_1.state.is_configured() and 
            self.config_panel_2.state.is_configured()):
            self.refresh_btn.setEnabled(True)
            self.export_btn.setEnabled(True)
            self.info_label.setText("✓ Оба файла настроены. Начинается умное сравнение...")
            # Используем QTimer чтобы UI успел обновиться перед тяжёлой операцией
            QtCore.QTimer.singleShot(100, self._load_comparison)

    def _load_comparison(self) -> None:
        """Загрузить умное сравнение"""
        try:
            self.info_label.setText("⏳ Выполняется анализ и сравнение данных...")
            QtWidgets.QApplication.processEvents()
            
            rows1 = self.config_panel_1.get_all_rows()
            rows2 = self.config_panel_2.get_all_rows()
            columns1 = self.config_panel_1.get_columns()
            columns2 = self.config_panel_2.get_columns()
            
            # Проверка что данные реально получены
            if not rows1 and not rows2:
                self.info_label.setText("⚠️ Оба файла пустые — нечего сравнивать")
                return
            
            if not columns1:
                self.info_label.setText("⚠️ В файле #1 не выбраны колонки")
                return
            
            if not columns2:
                self.info_label.setText("⚠️ В файле #2 не выбраны колонки")
                return
            
            self.comparison_table.load_comparison_smart(rows1, rows2, columns1, columns2)
            
            # Принудительно обновляем отображение таблицы
            self.comparison_table.resizeColumnsToContents()
            self.comparison_table.viewport().update()
            
            file1_name = self.config_panel_1.get_file_name()
            file2_name = self.config_panel_2.get_file_name()
            
            total_rows_in_table = self.comparison_table.rowCount()
            
            self.info_label.setText(
                f"✓ Сравнение завершено | {file1_name}: {len(rows1)} строк | "
                f"{file2_name}: {len(rows2)} строк | "
                f"Итого в таблице: {total_rows_in_table} строк"
            )
        except ValueError as ve:
            self.info_label.setText(f"⚠️ {str(ve)}")
        except Exception as e:
            self.info_label.setText(f"✗ Ошибка: {str(e)}")
            import traceback
            traceback.print_exc()

    def _export_report(self) -> None:
        """Экспортировать отчет"""
        if self.comparison_table.rowCount() == 0:
            QtWidgets.QMessageBox.warning(
                self, "Нет данных",
                "Сначала выполните сравнение файлов"
            )
            return
        
        file_path, _ = QtWidgets.QFileDialog.getSaveFileName(
            self,
            "Сохранить отчет сравнения",
            "comparison_report.xlsx",
            "Excel (*.xlsx);;Все файлы (*.*)"
        )
        
        if not file_path:
            return
        
        if not file_path.endswith('.xlsx'):
            file_path += '.xlsx'
        
        try:
            file1_name = self.config_panel_1.get_file_name()
            file2_name = self.config_panel_2.get_file_name()
            
            self.comparison_table.export_to_excel(file_path, file1_name, file2_name)
            
            QtWidgets.QMessageBox.information(
                self, "Успех",
                f"✓ Отчет успешно сохранен:\n{file_path}"
            )
        except Exception as e:
            QtWidgets.QMessageBox.critical(
                self, "Ошибка при экспорте",
                f"✗ Ошибка при сохранении файла:\n{str(e)}"
            )


# ============================================================================
# MAIN
# ============================================================================
def main() -> None:
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication(sys.argv)
    window = MainWindow()
    window.show()
    sys.exit(app.exec())


if __name__ == "__main__":
    main()
