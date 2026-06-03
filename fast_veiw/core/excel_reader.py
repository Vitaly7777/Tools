import warnings
from pathlib import Path
from typing import List, Tuple, Any, Optional, Iterator
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet


class ExcelReader:
    def __init__(self, file_path: Path):
        self.file_path = file_path
        self._workbook = None

    def __enter__(self):
        with warnings.catch_warnings():
            warnings.simplefilter("ignore")
            self._workbook = load_workbook(
                self.file_path, read_only=True, data_only=True
            )
        return self

    def __exit__(self, exc_type, exc_val, exc_tb):
        if self._workbook:
            self._workbook.close()

    @property
    def sheet_names(self) -> List[str]:
        return self._workbook.sheetnames

    def get_sheet_metadata(self, sheet_name: str) -> Tuple[int, int]:
        sheet = self._workbook[sheet_name]
        return sheet.max_row, sheet.max_column

    def read_rows(
        self,
        sheet_name: str,
        start_row: int = 1,
        max_rows: Optional[int] = None,
        stop_check=None,
    ) -> Tuple[List[Tuple[Any, ...]], int, int]:
        sheet = self._workbook[sheet_name]
        rows_data: List[Tuple[Any, ...]] = []
        effective = 0
        filled = 0

        end_row = None if max_rows is None else start_row + max_rows - 1
        iterator = sheet.iter_rows(
            min_row=start_row, max_row=end_row, values_only=True
        )

        for idx, row in enumerate(iterator, start=start_row):
            if stop_check and stop_check():
                break
            row_tuple = tuple(row)
            rows_data.append(row_tuple)
            if any(cell not in (None, "") for cell in row_tuple):
                effective += 1
                filled = idx

        return rows_data, effective, filled

    def calculate_metrics_with_stable_check(
        self,
        sheet_name: str,
        start_row: int,
        stable_threshold: int,
        stop_check=None,
    ) -> Tuple[int, int, int, bool]:
        sheet = self._workbook[sheet_name]
        effective = 0
        filled = 0
        last_filled = 0
        stable_counter = 0
        rows_read = 0
        is_stable = False

        iterator = sheet.iter_rows(min_row=start_row, values_only=True)

        for idx, row in enumerate(iterator, start=start_row):
            if stop_check and stop_check():
                break
            rows_read += 1
            if any(cell not in (None, "") for cell in row):
                effective += 1
                filled = idx
                if filled == last_filled:
                    stable_counter += 1
                else:
                    last_filled = filled
                    stable_counter = 0
            else:
                if filled == last_filled:
                    stable_counter += 1

            if stable_counter >= stable_threshold:
                is_stable = True
                break

        return effective, filled, rows_read, is_stable
        
    def get_active_sheet_name(self) -> str:
        return self._workbook.active.title    
    
    def iter_sheet_rows(self, sheet_name: str, min_row: int, max_row: Optional[int] = None):
        """
        Возвращает итератор по строкам листа.
        """
        sheet = self._workbook[sheet_name]
        return sheet.iter_rows(min_row=min_row, max_row=max_row, values_only=True)    
